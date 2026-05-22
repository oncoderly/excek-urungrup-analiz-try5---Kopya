import streamlit as st
import pandas as pd
import numpy as np
import hashlib
import plotly.express as px
import plotly.graph_objects as go
from html import escape
from io import BytesIO
from pathlib import Path
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
MEMORY_VARIANT_COL = "Analiz Varyanti"
MEMORY_VARIANT_KEY_COL = "Analiz Teklif Kodu"
MEMORY_DEVICE_INCLUDED_COL = "Cihazlar Dahil"
VARIANT_DEVICE_INCLUDED = "Cihaz Dahil"
VARIANT_DEVICE_EXCLUDED = "Cihaz Haric"
CIHAZLAR_GROUP_KEY = "CIHAZLAR"
CONS_COL_GROUP = "Urun Grubu"
CONS_COL_PAGE = "Sayfa"
CONS_COL_DESC = "Urun Aciklamasi"
CONS_COL_MATERIAL = "Malzeme Fiyati"
CONS_COL_LABOR = "Iscilik Fiyati"
CONS_COL_GGK = "GGK Fiyati"
CONS_COL_TOTAL = "Genel Toplam"
CONS_COL_COUNT = "Bulunan Kayit Sayisi"
CONS_COL_QTY = "Miktar"
CONS_COL_UNIT = "Birim"
CONS_COL_MATERIAL_PCT = "Malzeme %"
CONS_COL_LABOR_PCT = "Iscilik %"
CONS_COL_GGK_PCT = "GGK %"
CONS_COL_TOTAL_PCT = "Genel Toplam %"
CONS_COL_UNIT_SCOPE = "Birim Fiyat Kapsami"
CONS_COL_MATERIAL_UNIT_PRICE = "Malzeme Birim Fiyati"
CONS_COL_LABOR_UNIT_PRICE = "Iscilik Birim Fiyati"
CONS_COL_GGK_UNIT_PRICE = "GGK Birim Fiyati"
CONS_COL_TOTAL_UNIT_PRICE = "Genel Toplam Birim Fiyati"
CONS_COL_OFFER_CODE = "Teklif Kodu"
CONS_COL_OFFER_DATE = "Teklif Tarihi"
CONS_COL_SELECTED_SHEETS = "Secilen Sayfalar"
CONS_COL_SAVED_AT = "Kayit Zamani"
UNIT_PRICE_SCOPE_WITH_UNIT = "Urun Grubu + Birim"
UNIT_PRICE_SCOPE_GROUP_ONLY = "Sadece Urun Grubu"

PRICE_COLUMNS = ['Malzeme Fiyatı', 'İşçilik Fiyatı', 'GGK Fiyatı', 'Genel Toplam']

def format_currency_display(value):
    """Format numeric value as Turkish currency text for display without changing dtype."""
    if pd.isna(value):
        return ""
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return str(value)
    if numeric_value == 0:
        return "0,00 TL"
    return f"{numeric_value:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")

def format_percent_display(value):
    """Format numeric value as percent text for display without changing dtype."""
    if pd.isna(value):
        return ""
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{numeric_value:.2f}%"

def format_integer_display(value):
    """Format numeric value without decimal digits for display without changing dtype."""
    if pd.isna(value):
        return ""
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{numeric_value:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

def calculate_cumulative_percent_by_descending_value(dataframe, value_col, total_value=None):
    """Calculate cumulative percent for a metric's own descending order."""
    if dataframe is None or value_col not in dataframe.columns:
        return pd.Series(dtype="float64")

    values = pd.to_numeric(dataframe[value_col], errors="coerce").fillna(0.0)
    if total_value is None:
        total_value = values.sum()

    try:
        numeric_total = float(total_value)
    except (TypeError, ValueError):
        numeric_total = 0.0

    if numeric_total <= 0:
        return pd.Series(0.0, index=dataframe.index)

    ordered = pd.DataFrame({"_value": values}, index=dataframe.index).sort_values(
        "_value",
        ascending=False,
        kind="mergesort",
    )
    ordered["_cumulative_percent"] = ordered["_value"].cumsum() / numeric_total * 100
    ordered["_cumulative_percent"] = ordered.groupby("_value", sort=False)["_cumulative_percent"].transform("max")
    return ordered["_cumulative_percent"].round(2).reindex(dataframe.index).fillna(0.0)

def normalize_excel_header_name(value):
    """Normalize Turkish column names for Excel style matching."""
    text = "" if value is None else str(value)
    tr_chars = {
        "ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c",
        "İ": "I", "Ğ": "G", "Ü": "U", "Ş": "S", "Ö": "O", "Ç": "C",
    }
    for tr, en in tr_chars.items():
        text = text.replace(tr, en)
    return text.upper().strip()

def prepare_dataframe_for_excel_export(dataframe):
    """Convert percent columns to fraction for native Excel % formatting."""
    export_df = dataframe.copy()
    percent_cols = [col for col in export_df.columns if "%" in str(col)]
    for col in percent_cols:
        numeric_series = pd.to_numeric(export_df[col], errors="coerce")
        valid_values = numeric_series.dropna()
        if valid_values.empty:
            continue
        if valid_values.abs().max() > 1:
            numeric_series = numeric_series / 100.0
        export_df[col] = numeric_series
    return export_df

def style_excel_sheet(worksheet):
    """Apply Excel-friendly formatting, colors, and fonts to exported sheets."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_default_fill = PatternFill(fill_type="solid", fgColor="7C3AED")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11, color="1F2937")
    body_font_bold = Font(name="Calibri", size=11, bold=True, color="111827")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    zebra_even_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    zebra_odd_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    palette = {
        "malzeme": {"header": "10B981", "body": "DCFCE7"},
        "iscilik": {"header": "3B82F6", "body": "DBEAFE"},
        "ggk": {"header": "8B5CF6", "body": "E9D5FF"},
        "genel": {"header": "EF4444", "body": "FEE2E2"},
        "grup": {"header": "6B7280", "body": "F3F4F6"},
    }

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 26

    max_col = worksheet.max_column
    max_row = worksheet.max_row

    for col_idx in range(1, max_col + 1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_name = "" if header_cell.value is None else str(header_cell.value)
        normalized_header = normalize_excel_header_name(header_name)

        is_malzeme_col = "MALZEME" in normalized_header
        is_iscilik_col = "ISCILIK" in normalized_header
        is_ggk_col = "GGK" in normalized_header
        is_genel_col = ("GENEL TOPLAM" in normalized_header) or ("KUMULATIF GENEL" in normalized_header)
        is_group_col = "URUN GRUBU" in normalized_header
        is_percent_col = "%" in header_name
        is_currency_col = (
            any(token in normalized_header for token in ["MALZEME FIYATI", "ISCILIK FIYATI", "GGK FIYATI", "GENEL TOPLAM"])
            and not is_percent_col
        )
        is_integer_col = any(token in normalized_header for token in ["KAYIT SAYISI", "SIRA", "SATIR", "TOPLAM MIKTAR"])

        col_data_fill = None
        col_header_fill = header_default_fill
        if is_genel_col:
            col_header_fill = PatternFill(fill_type="solid", fgColor=palette["genel"]["header"])
            col_data_fill = PatternFill(fill_type="solid", fgColor=palette["genel"]["body"])
        elif is_malzeme_col:
            col_header_fill = PatternFill(fill_type="solid", fgColor=palette["malzeme"]["header"])
            col_data_fill = PatternFill(fill_type="solid", fgColor=palette["malzeme"]["body"])
        elif is_iscilik_col:
            col_header_fill = PatternFill(fill_type="solid", fgColor=palette["iscilik"]["header"])
            col_data_fill = PatternFill(fill_type="solid", fgColor=palette["iscilik"]["body"])
        elif is_ggk_col:
            col_header_fill = PatternFill(fill_type="solid", fgColor=palette["ggk"]["header"])
            col_data_fill = PatternFill(fill_type="solid", fgColor=palette["ggk"]["body"])
        elif is_group_col:
            col_header_fill = PatternFill(fill_type="solid", fgColor=palette["grup"]["header"])
            col_data_fill = PatternFill(fill_type="solid", fgColor=palette["grup"]["body"])

        header_cell.fill = col_header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = thin_border

        sample_last_row = min(max_row, 500)
        max_text_len = len(header_name)
        for row_idx in range(2, sample_last_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            if isinstance(cell_value, (int, float, np.integer, np.floating)):
                cell_text = f"{cell_value:,.2f}"
            else:
                cell_text = str(cell_value)
            max_text_len = max(max_text_len, len(cell_text))

        min_width = 12
        if is_currency_col:
            min_width = 16
        elif is_percent_col:
            min_width = 14
        elif is_group_col:
            min_width = 20
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_text_len + 2, min_width), 45)

        for row_idx in range(2, max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            worksheet.row_dimensions[row_idx].height = 20

            zebra_fill = zebra_even_fill if row_idx % 2 == 0 else zebra_odd_fill
            cell.fill = col_data_fill if col_data_fill is not None else zebra_fill
            cell.border = thin_border

            if is_group_col:
                cell.font = body_font_bold
            else:
                cell.font = body_font

            if is_percent_col:
                cell.number_format = "0.00%"
            elif is_currency_col:
                cell.number_format = '#,##0.00 "TL"'
            elif is_integer_col:
                cell.number_format = "#,##0"

            if is_currency_col or is_percent_col or is_integer_col:
                cell.alignment = right_alignment
            elif is_group_col or "URUN ACIKLAMASI" in normalized_header or "SAYFA" in normalized_header:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

    worksheet.auto_filter.ref = worksheet.dimensions

def index_to_excel_column(index):
    """Convert zero-based column index to Excel column letter (A, B, AA)."""
    if index is None or index < 0:
        return "?"

    result = ""
    number = index + 1
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def get_sheet_columns(excel_file, sheet_name):
    """Read sheet columns robustly; fall back to positional columns if header row is empty."""
    try:
        header_df = excel_file.parse(sheet_name=sheet_name, nrows=0)
        columns = list(header_df.columns)
    except Exception as exc:
        return [], str(exc)

    if columns:
        return columns, None

    try:
        sample_df = excel_file.parse(sheet_name=sheet_name, header=None, nrows=1)
        col_count = int(sample_df.shape[1])
        if col_count > 0:
            return [f"Unnamed: {i}" for i in range(col_count)], None
    except Exception:
        pass

    return [], None

def derive_offer_code_from_filename(file_name):
    """Derive offer code from uploaded file name."""
    stem = Path(str(file_name)).stem.strip()
    if not stem:
        stem = f"TEKLIF-{datetime.now().strftime('%Y%m%d-%H%M')}"
    return stem

def derive_offer_code_from_outputs(all_df, selected_sheets=None, fallback_name=None):
    """Prefer uploaded file name as offer code; fall back to deterministic content hash if needed."""
    if fallback_name:
        return derive_offer_code_from_filename(fallback_name)

    if all_df is None or all_df.empty:
        return derive_offer_code_from_filename(fallback_name)

    signature_columns = [
        col for col in [
            "Ürün Grubu",
            "Sayfa",
            "Ürün Açıklaması",
            "Malzeme Fiyatı",
            "İşçilik Fiyatı",
            "GGK Fiyatı",
            "Genel Toplam",
        ]
        if col in all_df.columns
    ]
    if not signature_columns:
        return derive_offer_code_from_filename(fallback_name)

    signature_df = all_df[signature_columns].copy()
    for col in signature_columns:
        if col in PRICE_COLUMNS:
            signature_df[col] = pd.to_numeric(signature_df[col], errors="coerce").fillna(0).round(4)
        else:
            signature_df[col] = signature_df[col].fillna("").astype(str).str.strip().str.lower()

    signature_df = signature_df.sort_values(signature_columns).reset_index(drop=True)
    row_hashes = pd.util.hash_pandas_object(signature_df, index=False).to_numpy(dtype="uint64")
    hasher = hashlib.sha1()
    hasher.update(row_hashes.tobytes())

    if selected_sheets:
        normalized_sheets = "|".join(sorted(str(sheet).strip().lower() for sheet in selected_sheets))
        hasher.update(normalized_sheets.encode("utf-8"))

    digest = hasher.hexdigest()[:12].upper()
    return f"EXCEL-{digest}"

def get_offer_memory_paths():
    base_path = Path(__file__).resolve().parent
    return (
        base_path / "teklif_hafiza_grup.csv",
        base_path / "teklif_hafiza_urun.csv",
        base_path / "teklif_hafiza_sistem.csv",
    )

def get_unit_price_memory_path():
    return Path(__file__).resolve().parent / "teklif_hafiza_birim_fiyat.csv"

def load_memory_dataframe(file_path):
    if not Path(file_path).exists():
        return pd.DataFrame()
    try:
        df = pd.read_csv(file_path)
    except Exception:
        return pd.DataFrame()

    for date_col in ["Teklif Tarihi", "Kayıt Zamanı", CONS_COL_OFFER_DATE, CONS_COL_SAVED_AT]:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    return df

def upsert_memory_dataframe(file_path, dataframe, key_columns):
    def build_row_keys(df, columns):
        key_frame = df.loc[:, columns].copy()
        if isinstance(key_frame, pd.Series):
            key_frame = key_frame.to_frame()
        key_frame = key_frame.fillna("").astype(str)
        return key_frame.apply(lambda row: "|".join(row.tolist()), axis=1)

    dataframe = dataframe.copy()
    key_columns = list(dict.fromkeys(key_columns))
    existing_df = load_memory_dataframe(file_path)
    if existing_df.empty:
        final_df = dataframe.copy()
    else:
        # Aynı dosya adıyla (Teklif Kodu) tekrar yüklenirse eski kaydı komple silip son yüklemeyi bırak.
        existing_has_variant_key = MEMORY_VARIANT_KEY_COL in existing_df.columns
        if MEMORY_VARIANT_KEY_COL in dataframe.columns and existing_has_variant_key:
            incoming_offer_keys = set(
                dataframe[MEMORY_VARIANT_KEY_COL].astype(str).str.strip().replace("", np.nan).dropna().tolist()
            )
            if incoming_offer_keys:
                existing_df = existing_df[
                    ~existing_df[MEMORY_VARIANT_KEY_COL].astype(str).str.strip().isin(incoming_offer_keys)
                ].copy()
        elif "Teklif Kodu" in dataframe.columns and "Teklif Kodu" in existing_df.columns:
            incoming_offer_codes = set(
                dataframe["Teklif Kodu"].astype(str).str.strip().replace("", np.nan).dropna().tolist()
            )
            if incoming_offer_codes:
                existing_df = existing_df[
                    ~existing_df["Teklif Kodu"].astype(str).str.strip().isin(incoming_offer_codes)
                ].copy()

        for key_col in key_columns:
            if key_col not in existing_df.columns:
                existing_df[key_col] = ""
            if key_col not in dataframe.columns:
                dataframe[key_col] = ""
        existing_keys = build_row_keys(existing_df, key_columns)
        incoming_keys = build_row_keys(dataframe, key_columns)
        final_df = pd.concat(
            [
                existing_df[~existing_keys.isin(set(incoming_keys.tolist()))],
                dataframe
            ],
            ignore_index=True
        )
    final_df.to_csv(file_path, index=False, encoding="utf-8-sig")

def apply_context_filter(memory_df, bina_tipi, ana_cihaz_durumu, is_konumu, bina_alt_tipi=None):
    if memory_df.empty:
        return memory_df.copy()
    filtered_df = memory_df.copy()
    for col_name, selected_value in [
        ("Bina Tipi", bina_tipi),
        ("Bina Alt Tipi", bina_alt_tipi),
        ("Ana Cihaz Durumu", ana_cihaz_durumu),
        ("İş Konumu", is_konumu),
    ]:
        if selected_value is not None and col_name in filtered_df.columns:
            if col_name == "Bina Alt Tipi":
                col_text = filtered_df[col_name].astype(str).str.strip()
                filtered_df = filtered_df[(col_text == str(selected_value)) | (col_text == "") | (col_text.str.lower() == "nan")]
            else:
                filtered_df = filtered_df[filtered_df[col_name] == selected_value]
    return filtered_df

def normalize_sheet_name(name):
    return turkce_ascii(str(name)).upper().strip()

def get_cost_sheet_names(sheet_names):
    excluded = {"ISKONTOLAR", "GENEL GIDER ANALIZ"}
    return [sheet for sheet in sheet_names if normalize_sheet_name(sheet) not in excluded]

def map_sheet_names_by_normalized(sheet_names):
    mapping = {}
    for sheet in sheet_names:
        normalized = normalize_sheet_name(sheet)
        if normalized not in mapping:
            mapping[normalized] = sheet
    return mapping

def read_product_groups_from_iskontolar(
    excel_source,
    iskontolar_sheet_name,
    group_col_index,
    start_row,
    end_row,
):
    if not iskontolar_sheet_name:
        return []

    try:
        iskontolar_df = pd.read_excel(excel_source, sheet_name=iskontolar_sheet_name)
    except Exception:
        return []

    if group_col_index is None or group_col_index < 0:
        return []
    if len(iskontolar_df.columns) <= group_col_index:
        return []

    row_start = max(min(int(start_row), int(end_row)) - 1, 0)
    row_end = max(int(start_row), int(end_row))

    groups = []
    for i in range(row_start, row_end):
        if i >= len(iskontolar_df):
            continue
        value = iskontolar_df.iloc[i, group_col_index]
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            groups.append(text)
    return list(dict.fromkeys(groups))

def build_cost_dataframe_from_sheets(
    excel_source,
    sheet_names,
    allowed_product_groups,
    data_group_col_index,
    urun_aciklama_col_index,
    malzeme_col_index,
    iscilik_col_index,
    ggk_col_index,
    genel_toplam_col_index,
):
    all_data = []
    selected_indices = [
        data_group_col_index,
        urun_aciklama_col_index,
        malzeme_col_index,
        iscilik_col_index,
        ggk_col_index,
        genel_toplam_col_index,
    ]
    selected_indices = [idx for idx in selected_indices if idx is not None]
    if not selected_indices:
        return pd.DataFrame()

    required_max_col = max(selected_indices)
    allowed_set = set(allowed_product_groups or [])

    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(excel_source, sheet_name=sheet_name)
        except Exception:
            continue

        if data_group_col_index is None or len(df.columns) <= required_max_col:
            continue

        for idx, row in df.iterrows():
            try:
                if pd.isna(row.iloc[data_group_col_index]):
                    continue
                product_group = str(row.iloc[data_group_col_index]).strip()
                if not product_group or product_group.lower() == "nan":
                    continue
                if allowed_set and product_group not in allowed_set:
                    continue

                all_data.append(
                    {
                        "Ürün Grubu": product_group,
                        "Sayfa": sheet_name,
                        "Satır": idx + 1,
                        "Ürün Açıklaması": str(row.iloc[urun_aciklama_col_index]).strip()
                        if len(row) > urun_aciklama_col_index and pd.notna(row.iloc[urun_aciklama_col_index])
                        else "",
                        "Malzeme Fiyatı": float(row.iloc[malzeme_col_index])
                        if len(row) > malzeme_col_index and pd.notna(row.iloc[malzeme_col_index])
                        else 0.0,
                        "İşçilik Fiyatı": float(row.iloc[iscilik_col_index])
                        if len(row) > iscilik_col_index and pd.notna(row.iloc[iscilik_col_index])
                        else 0.0,
                        "GGK Fiyatı": float(row.iloc[ggk_col_index])
                        if len(row) > ggk_col_index and pd.notna(row.iloc[ggk_col_index])
                        else 0.0,
                        "Genel Toplam": float(row.iloc[genel_toplam_col_index])
                        if len(row) > genel_toplam_col_index and pd.notna(row.iloc[genel_toplam_col_index])
                        else 0.0,
                    }
                )
            except Exception:
                continue

    return pd.DataFrame(all_data)

def render_price_revision_compare_module(
    new_file,
    old_file,
    selected_new_cost_sheets,
    iskontolar_group_col_index,
    iskontolar_start_row,
    iskontolar_end_row,
    data_group_col_index,
    urun_aciklama_col_index,
    malzeme_col_index,
    iscilik_col_index,
    ggk_col_index,
    genel_toplam_col_index,
):
    render_section_heading("Fiyat Revizyon Kıyas", icon="")
    st.caption("Yeni ve eski fiyat çalışması dosyaları satır/sayfa/ürün grubu bazında karşılaştırılır.")

    try:
        new_excel = pd.ExcelFile(new_file)
        old_excel = pd.ExcelFile(old_file)
    except Exception as exc:
        st.error(f"Dosyalar okunamadı: {str(exc)}")
        return

    new_sheet_map = map_sheet_names_by_normalized(new_excel.sheet_names)
    old_sheet_map = map_sheet_names_by_normalized(old_excel.sheet_names)

    requested_new_sheets = selected_new_cost_sheets if selected_new_cost_sheets else get_cost_sheet_names(new_excel.sheet_names)
    sheet_pairs = []
    missing_old_sheets = []
    for new_sheet in requested_new_sheets:
        normalized = normalize_sheet_name(new_sheet)
        if normalized in ("ISKONTOLAR", "GENEL GIDER ANALIZ"):
            continue
        old_sheet = old_sheet_map.get(normalized)
        if old_sheet:
            sheet_pairs.append((new_sheet, old_sheet))
        else:
            missing_old_sheets.append(new_sheet)

    if missing_old_sheets:
        st.warning(
            "Eski dosyada bulunamayan sayfalar kıyasa dahil edilmedi: "
            + ", ".join(missing_old_sheets[:8])
            + (" ..." if len(missing_old_sheets) > 8 else "")
        )

    if not sheet_pairs:
        st.error("Karşılaştırılacak ortak maliyet sayfası bulunamadı.")
        return

    new_iskontolar_sheet = new_sheet_map.get("ISKONTOLAR")
    old_iskontolar_sheet = old_sheet_map.get("ISKONTOLAR")

    new_groups = read_product_groups_from_iskontolar(
        new_file,
        new_iskontolar_sheet,
        iskontolar_group_col_index,
        iskontolar_start_row,
        iskontolar_end_row,
    )
    old_groups = read_product_groups_from_iskontolar(
        old_file,
        old_iskontolar_sheet,
        iskontolar_group_col_index,
        iskontolar_start_row,
        iskontolar_end_row,
    )

    if not new_groups:
        st.info("Yeni dosyada ISKONTOLAR ürün grupları bulunamadı. Maliyet sayfalarındaki tüm ürün grupları kullanılacak.")
    if not old_groups:
        st.info("Eski dosyada ISKONTOLAR ürün grupları bulunamadı. Maliyet sayfalarındaki tüm ürün grupları kullanılacak.")

    new_df = build_cost_dataframe_from_sheets(
        new_file,
        [new_sheet for new_sheet, _ in sheet_pairs],
        new_groups,
        data_group_col_index,
        urun_aciklama_col_index,
        malzeme_col_index,
        iscilik_col_index,
        ggk_col_index,
        genel_toplam_col_index,
    )
    old_df = build_cost_dataframe_from_sheets(
        old_file,
        [old_sheet for _, old_sheet in sheet_pairs],
        old_groups,
        data_group_col_index,
        urun_aciklama_col_index,
        malzeme_col_index,
        iscilik_col_index,
        ggk_col_index,
        genel_toplam_col_index,
    )

    if not new_df.empty:
        new_df["Sayfa"] = new_df["Sayfa"].apply(normalize_sheet_name)
    if not old_df.empty:
        old_df["Sayfa"] = old_df["Sayfa"].apply(normalize_sheet_name)

    if new_df.empty and old_df.empty:
        st.error("Her iki dosyada da karşılaştırılabilir maliyet satırı bulunamadı.")
        return

    key_cols = ["Ürün Grubu", "Sayfa", "Ürün Açıklaması"]
    cost_cols = ["Malzeme Fiyatı", "İşçilik Fiyatı", "GGK Fiyatı", "Genel Toplam"]

    def aggregate_for_compare(dataframe, prefix):
        output_cols = key_cols + [f"{prefix} Excel Satır No"] + [f"{prefix} {col}" for col in cost_cols]
        if dataframe.empty:
            return pd.DataFrame(columns=output_cols)
        aggregated = dataframe.groupby(key_cols, as_index=False).agg(
            {
                "Satır": "min",
                "Malzeme Fiyatı": "sum",
                "İşçilik Fiyatı": "sum",
                "GGK Fiyatı": "sum",
                "Genel Toplam": "sum",
            }
        )
        return aggregated.rename(
            columns={
                "Satır": f"{prefix} Excel Satır No",
                "Malzeme Fiyatı": f"{prefix} Malzeme Fiyatı",
                "İşçilik Fiyatı": f"{prefix} İşçilik Fiyatı",
                "GGK Fiyatı": f"{prefix} GGK Fiyatı",
                "Genel Toplam": f"{prefix} Genel Toplam",
            }
        )

    old_agg = aggregate_for_compare(old_df, "Eski")
    new_agg = aggregate_for_compare(new_df, "Yeni")

    compare_df = old_agg.merge(new_agg, on=key_cols, how="outer")
    for col in [f"Eski {name}" for name in cost_cols] + [f"Yeni {name}" for name in cost_cols]:
        compare_df[col] = pd.to_numeric(compare_df[col], errors="coerce").fillna(0.0)
    compare_df["Eski Excel Satır No"] = pd.to_numeric(compare_df["Eski Excel Satır No"], errors="coerce")
    compare_df["Yeni Excel Satır No"] = pd.to_numeric(compare_df["Yeni Excel Satır No"], errors="coerce")
    compare_df["Excel Satır No"] = (
        compare_df["Yeni Excel Satır No"]
        .combine_first(compare_df["Eski Excel Satır No"])
        .round()
        .astype("Int64")
    )

    compare_df["Genel Toplam Farkı"] = (compare_df["Yeni Genel Toplam"] - compare_df["Eski Genel Toplam"]).round(2)
    compare_df["İndirim Tutarı"] = np.where(compare_df["Genel Toplam Farkı"] < 0, -compare_df["Genel Toplam Farkı"], 0).round(2)
    compare_df["Zam Tutarı"] = np.where(compare_df["Genel Toplam Farkı"] > 0, compare_df["Genel Toplam Farkı"], 0).round(2)
    compare_df["Genel İndirim Tutarı"] = compare_df["İndirim Tutarı"]
    compare_df["Genel Zam Tutarı"] = compare_df["Zam Tutarı"]

    compare_df["Malzeme Farkı"] = (compare_df["Yeni Malzeme Fiyatı"] - compare_df["Eski Malzeme Fiyatı"]).round(2)
    compare_df["Malzeme İndirim Tutarı"] = np.where(compare_df["Malzeme Farkı"] < 0, -compare_df["Malzeme Farkı"], 0).round(2)
    compare_df["Malzeme Zam Tutarı"] = np.where(compare_df["Malzeme Farkı"] > 0, compare_df["Malzeme Farkı"], 0).round(2)

    compare_df["İşçilik Farkı"] = (compare_df["Yeni İşçilik Fiyatı"] - compare_df["Eski İşçilik Fiyatı"]).round(2)
    compare_df["İşçilik İndirim Tutarı"] = np.where(compare_df["İşçilik Farkı"] < 0, -compare_df["İşçilik Farkı"], 0).round(2)
    compare_df["İşçilik Zam Tutarı"] = np.where(compare_df["İşçilik Farkı"] > 0, compare_df["İşçilik Farkı"], 0).round(2)

    def pct_change(old_series, new_series):
        return np.where(
            old_series > 0,
            ((new_series - old_series) / old_series * 100),
            np.where(new_series > 0, 100.0, 0.0),
        ).round(2)

    compare_df["Genel Değişim %"] = pct_change(compare_df["Eski Genel Toplam"], compare_df["Yeni Genel Toplam"])
    compare_df["Değişim %"] = compare_df["Genel Değişim %"]
    compare_df["Malzeme Değişim %"] = pct_change(compare_df["Eski Malzeme Fiyatı"], compare_df["Yeni Malzeme Fiyatı"])
    compare_df["İşçilik Değişim %"] = pct_change(compare_df["Eski İşçilik Fiyatı"], compare_df["Yeni İşçilik Fiyatı"])

    def detect_change_type(row):
        old_val = float(row["Eski Genel Toplam"])
        new_val = float(row["Yeni Genel Toplam"])
        if old_val == 0 and new_val > 0:
            return "Yeni Kalem"
        if old_val > 0 and new_val == 0:
            return "Kaldırılan Kalem"
        if new_val < old_val:
            return "İndirim"
        if new_val > old_val:
            return "Zam"
        return "Değişmedi"

    compare_df["Durum"] = compare_df.apply(detect_change_type, axis=1)

    total_old = float(compare_df["Eski Genel Toplam"].sum())
    total_new = float(compare_df["Yeni Genel Toplam"].sum())
    net_change = total_new - total_old
    total_discount = float(compare_df["İndirim Tutarı"].sum())
    total_increase = float(compare_df["Zam Tutarı"].sum())
    net_change_pct = (net_change / total_old * 100) if total_old > 0 else 0.0

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Eski Toplam", format_currency_display(total_old))
    with col2:
        st.metric("Yeni Toplam", format_currency_display(total_new))
    with col3:
        st.metric("Net Fark", format_currency_display(net_change), delta=f"{net_change_pct:.2f}%")
    with col4:
        st.metric("Toplam İndirim", format_currency_display(total_discount), delta=f"Zam: {format_currency_display(total_increase)}")

    status_options = ["İndirim", "Zam", "Yeni Kalem", "Kaldırılan Kalem", "Değişmedi"]
    selected_statuses = st.multiselect(
        "Durum filtresi",
        options=status_options,
        default=["İndirim", "Zam", "Yeni Kalem", "Kaldırılan Kalem"],
        key="price_compare_status_filter",
    )

    filtered_df = compare_df[compare_df["Durum"].isin(selected_statuses)].copy()
    filtered_df = filtered_df.sort_values("Genel Toplam Farkı", key=lambda s: s.abs(), ascending=False)

    st.markdown("### Ürün/Satır Bazında Fiyat Değişimi")
    if filtered_df.empty:
        st.info("Seçilen filtreye uygun kayıt bulunamadı.")
    else:
        detail_cols = [
            "Durum",
            "Ürün Grubu",
            "Sayfa",
            "Excel Satır No",
            "Ürün Açıklaması",
            "Eski Genel Toplam",
            "Yeni Genel Toplam",
            "Genel Toplam Farkı",
            "Genel Değişim %",
            "Genel İndirim Tutarı",
            "Genel Zam Tutarı",
            "Eski Malzeme Fiyatı",
            "Yeni Malzeme Fiyatı",
            "Malzeme Farkı",
            "Malzeme Değişim %",
            "Malzeme İndirim Tutarı",
            "Malzeme Zam Tutarı",
            "Eski İşçilik Fiyatı",
            "Yeni İşçilik Fiyatı",
            "İşçilik Farkı",
            "İşçilik Değişim %",
            "İşçilik İndirim Tutarı",
            "İşçilik Zam Tutarı",
            "Eski GGK Fiyatı",
            "Yeni GGK Fiyatı",
        ]
        detail_currency_cols = [
            "Eski Genel Toplam",
            "Yeni Genel Toplam",
            "Genel Toplam Farkı",
            "Genel İndirim Tutarı",
            "Genel Zam Tutarı",
            "Eski Malzeme Fiyatı",
            "Yeni Malzeme Fiyatı",
            "Malzeme Farkı",
            "Malzeme İndirim Tutarı",
            "Malzeme Zam Tutarı",
            "Eski İşçilik Fiyatı",
            "Yeni İşçilik Fiyatı",
            "İşçilik Farkı",
            "İşçilik İndirim Tutarı",
            "İşçilik Zam Tutarı",
            "Eski GGK Fiyatı",
            "Yeni GGK Fiyatı",
        ]
        create_sortable_numeric_table(
            filtered_df[detail_cols],
            height=520,
            currency_cols=detail_currency_cols,
            percent_cols=["Genel Değişim %", "Malzeme Değişim %", "İşçilik Değişim %"],
        )

    st.markdown("### Ürün Grubu Bazında Özet Kıyas")
    group_summary = (
        compare_df.groupby("Ürün Grubu", as_index=False)[
            [
                "Eski Genel Toplam",
                "Yeni Genel Toplam",
                "Genel Toplam Farkı",
                "İndirim Tutarı",
                "Zam Tutarı",
                "Eski Malzeme Fiyatı",
                "Yeni Malzeme Fiyatı",
                "Eski İşçilik Fiyatı",
                "Yeni İşçilik Fiyatı",
            ]
        ]
        .sum()
        .sort_values("Genel Toplam Farkı", key=lambda s: s.abs(), ascending=False)
    )
    group_summary["Genel Toplam Farkı %"] = pct_change(group_summary["Eski Genel Toplam"], group_summary["Yeni Genel Toplam"])
    group_summary["Malzeme Değişim %"] = pct_change(group_summary["Eski Malzeme Fiyatı"], group_summary["Yeni Malzeme Fiyatı"])
    group_summary["İşçilik Değişim %"] = pct_change(group_summary["Eski İşçilik Fiyatı"], group_summary["Yeni İşçilik Fiyatı"])

    if group_summary.empty:
        st.info("Ürün grubu bazında kıyaslanacak veri bulunamadı.")
    else:
        group_summary_display = group_summary[
            [
                "Ürün Grubu",
                "Eski Genel Toplam",
                "Yeni Genel Toplam",
                "Genel Toplam Farkı",
                "Genel Toplam Farkı %",
                "İndirim Tutarı",
                "Zam Tutarı",
                "Eski Malzeme Fiyatı",
                "Yeni Malzeme Fiyatı",
                "Malzeme Değişim %",
                "Eski İşçilik Fiyatı",
                "Yeni İşçilik Fiyatı",
                "İşçilik Değişim %",
            ]
        ]
        create_sortable_numeric_table(
            group_summary_display,
            height=380,
            currency_cols=[
                "Eski Genel Toplam",
                "Yeni Genel Toplam",
                "Genel Toplam Farkı",
                "İndirim Tutarı",
                "Zam Tutarı",
                "Eski Malzeme Fiyatı",
                "Yeni Malzeme Fiyatı",
                "Eski İşçilik Fiyatı",
                "Yeni İşçilik Fiyatı",
            ],
            percent_cols=["Genel Toplam Farkı %", "Malzeme Değişim %", "İşçilik Değişim %"],
        )
SYSTEM_TYPE_DEFINITIONS = {
    "HVAC SİSTEM TİPİ": [
        {"tip": "Chiller + Klima Santrali (Merkezi Sistem)", "keywords": ["chiller", "klima santrali", "ahu", "air handling unit"]},
        {"tip": "VRF / VRV Sistem", "keywords": ["vrf", "vrv"]},
        {"tip": "Rooftop Sistem", "keywords": ["rooftop", "paket klima"]},
        {"tip": "Fancoil + Kazan", "keywords": ["fancoil", "fan coil", "kazan", "boiler"]},
        {"tip": "Isı Pompası Sistemi", "keywords": ["isi pompasi", "ısı pompası", "heat pump"]},
        {"tip": "District Heating / Cooling bağlantılı", "keywords": ["district heating", "district cooling", "bolgesel isitma", "bölgesel ısıtma", "bolgesel sogutma", "bölgesel soğutma"]},
        {"tip": "Precision Cooling (Data Center tipi)", "keywords": ["precision cooling", "close control", "datacenter cooling", "data center cooling"]},
    ],
    "HAVALANDIRMA SİSTEM TİPİ": [
        {"tip": "Konfor havalandırma", "keywords": ["konfor havalandirma", "konfor havalandırma"]},
        {"tip": "Hijyenik havalandırma (Hastane Class)", "keywords": ["hijyenik havalandirma", "hijyenik havalandırma", "hepa", "ameliyathane havalandirma"]},
        {"tip": "Endüstriyel havalandırma", "keywords": ["endustriyel havalandirma", "endüstriyel havalandırma"]},
        {"tip": "Duman tahliye sistemi", "keywords": ["duman tahliye", "smoke exhaust"]},
        {"tip": "Basınçlandırma sistemi", "keywords": ["basinclandirma", "basınçlandırma", "pressure fan"]},
        {"tip": "Mutfak egzoz sistemi (Restaurant/Cafe için kritik)", "keywords": ["mutfak egzoz", "kitchen exhaust", "davlumbaz egzoz"]},
    ],
    "YANGIN TESİSATI SİSTEM TİPİ": [
        {"tip": "Sprinkler sistemi", "keywords": ["sprinkler"]},
        {"tip": "Kuru sistem", "keywords": ["kuru sistem", "dry pipe"]},
        {"tip": "Köpüklü sistem", "keywords": ["kopuklu", "köpüklü", "foam system"]},
        {"tip": "Yangın dolap hattı", "keywords": ["yangin dolap", "yangın dolap"]},
        {"tip": "Hidrant hattı", "keywords": ["hidrant"]},
        {"tip": "Yangın pompa istasyonu", "keywords": ["yangin pompa", "yangın pompa", "fire pump"]},
        {"tip": "Gazlı söndürme (FM200, Novec)", "keywords": ["fm200", "novec", "gazli sondurme", "gazlı söndürme"]},
        {"tip": "Davlumbaz içi söndürme (Restaurant için önemli)", "keywords": ["davlumbaz ici sondurme", "davlumbaz içi söndürme", "hood suppression"]},
    ],
    "SIHHİ TESİSAT SİSTEM TİPİ": [
        {"tip": "Standart temiz su + pis su", "keywords": ["temiz su", "pis su", "atik su", "atık su"]},
        {"tip": "Sıcak su sirkülasyon sistemi", "keywords": ["sicak su sirkulasyon", "sıcak su sirkülasyon"]},
        {"tip": "Gri su sistemi", "keywords": ["gri su", "gray water"]},
        {"tip": "Yağ tutucu sistemi (Restaurant için kritik)", "keywords": ["yag tutucu", "yağ tutucu", "grease trap"]},
        {"tip": "Atık su arıtma", "keywords": ["atik su aritma", "atık su arıtma"]},
        {"tip": "Hidrofor sistemi", "keywords": ["hidrofor", "booster pump"]},
    ],
    "PROSES TESİSATI (Varsa)": [
        {"tip": "Buhar hattı", "keywords": ["buhar hatti", "buhar hattı", "steam line"]},
        {"tip": "Basınçlı hava", "keywords": ["basincli hava", "basınçlı hava", "compressed air"]},
        {"tip": "Proses suyu", "keywords": ["proses suyu", "process water"]},
        {"tip": "Chilled water özel proses hattı", "keywords": ["chilled water", "ozel proses hatti", "özel proses hattı"]},
        {"tip": "Kimyasal hatlar", "keywords": ["kimyasal hat", "chemical line"]},
        {"tip": "Medikal gaz sistemi (Hastane)", "keywords": ["medikal gaz", "medical gas", "medikla gaz"]},
    ],
    "OTOMASYON VE KONTROL": [
        {"tip": "BMS (Building Management System)", "keywords": ["bms", "building management system"]},
        {"tip": "Otomasyonlu HVAC", "keywords": ["otomasyonlu hvac", "hvac otomasyon"]},
        {"tip": "Enerji izleme sistemi", "keywords": ["enerji izleme", "energy monitoring"]},
        {"tip": "SCADA bağlantısı", "keywords": ["scada"]},
        {"tip": "Zon kontrollü sistem", "keywords": ["zon kontrollu", "zon kontrollü", "zone control"]},
    ],
}

BUILDING_TYPE_TREE = {
    "Endüstriyel Tesisler": [
        "Fabrikalar",
        "Üretim tesisleri",
        "Lojistik ve depo merkezleri",
        "Enerji üretim tesisleri",
        "Veri merkezleri (Data Center)",
    ],
    "Ticari ve Karma Kullanımlı Yapılar": [
        "A Sınıfı ofis binaları",
        "İş kuleleri",
        "Karma projeler (ofis + konut + ticari alan)",
        "Rezidans kompleksleri",
        "Restaurant & Cafe",
    ],
    "Sağlık Yapıları": [
        "Şehir hastaneleri",
        "Özel hastane kompleksleri",
        "Üniversite hastaneleri",
        "Sağlık kampüsleri",
    ],
    "Alışveriş ve Büyük Perakende Yapıları": [
        "Alışveriş merkezleri (AVM)",
        "Outlet ve ticari yaşam merkezleri",
    ],
    "Otel ve Turizm Yapıları": [
        "4-5 yıldızlı şehir otelleri",
        "Resort ve tatil köyleri",
        "Kongre otelleri",
    ],
    "Eğitim ve Kampüs Yapıları": [
        "Üniversite kampüsleri",
        "Fakülte ve araştırma binaları",
    ],
    "Kamu ve İdari Yapılar": [
        "Bakanlık ve kamu hizmet binaları",
        "Adliye sarayları",
        "Büyük ölçekli idari kompleksler",
    ],
    "Spor ve Kültür Yapıları": [
        "Stadyumlar",
        "Spor kompleksleri",
        "Kongre ve fuar merkezleri",
    ],
}

def normalize_text_for_match(value):
    text = str(value)
    tr_chars = {
        "ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c",
        "İ": "I", "Ğ": "G", "Ü": "U", "Ş": "S", "Ö": "O", "Ç": "C",
    }
    for tr, en in tr_chars.items():
        text = text.replace(tr, en)
    return text.lower().strip()

def get_all_system_type_rows():
    rows = []
    for category, definitions in SYSTEM_TYPE_DEFINITIONS.items():
        for definition in definitions:
            system_type = definition["tip"]
            rows.append(
                {
                    "Sistem Kategori": category,
                    "Sistem Tipi": system_type,
                    "Sistem Etiket": f"{category} / {system_type}",
                    "keywords": [normalize_text_for_match(k) for k in definition.get("keywords", [])],
                }
            )
    return rows

SYSTEM_TYPE_ROWS = get_all_system_type_rows()

def detect_system_classification(row):
    combined_text = " ".join(
        [
            str(row.get("Ürün Açıklaması", "")),
            str(row.get("Ürün Grubu", "")),
            str(row.get("Sayfa", "")),
        ]
    )
    normalized_text = normalize_text_for_match(combined_text)
    for entry in SYSTEM_TYPE_ROWS:
        if any(keyword in normalized_text for keyword in entry["keywords"]):
            return entry["Sistem Kategori"], entry["Sistem Tipi"], entry["Sistem Etiket"]
    return None, None, None

def build_system_distribution(all_df):
    system_df = all_df.copy()
    classifications = system_df.apply(detect_system_classification, axis=1, result_type="expand")
    classifications.columns = ["Sistem Kategori", "Sistem Tipi", "Sistem Etiket"]
    system_df = pd.concat([system_df, classifications], axis=1)
    system_df = system_df[system_df["Sistem Etiket"].notna()].copy()

    total_quote = pd.to_numeric(all_df["Genel Toplam"], errors="coerce").fillna(0).sum()
    all_types_df = pd.DataFrame(SYSTEM_TYPE_ROWS)[
        ["Sistem Kategori", "Sistem Tipi", "Sistem Etiket"]
    ]

    if system_df.empty:
        distribution = all_types_df.copy()
        distribution["Sistem Toplam"] = 0.0
        distribution["Sistem %"] = 0.0
        distribution["Sistem"] = distribution["Sistem Etiket"]
        return distribution

    grouped = (
        system_df.groupby(["Sistem Kategori", "Sistem Tipi", "Sistem Etiket"], as_index=False)["Genel Toplam"]
        .sum()
        .rename(columns={"Genel Toplam": "Sistem Toplam"})
    )
    grouped = all_types_df.merge(grouped, on=["Sistem Kategori", "Sistem Tipi", "Sistem Etiket"], how="left")
    grouped["Sistem Toplam"] = grouped["Sistem Toplam"].fillna(0.0)
    if total_quote > 0:
        grouped["Sistem %"] = (grouped["Sistem Toplam"] / total_quote * 100).round(2)
    else:
        grouped["Sistem %"] = 0.0
    grouped["Sistem"] = grouped["Sistem Etiket"]
    return grouped

def is_cihazlar_group(group_name):
    return turkce_ascii(str(group_name)).upper().strip() == CIHAZLAR_GROUP_KEY

def build_variant_offer_key(offer_code, variant_name):
    offer_text = str(offer_code).strip()
    if not offer_text:
        return str(variant_name).strip()
    return f"{offer_text} | {variant_name}"

def normalize_consistency_header(value):
    return normalize_excel_header_name(value).replace(" ", "")

def resolve_consistency_column(columns, *candidate_names):
    normalized_map = {}
    for column in columns:
        normalized_key = normalize_consistency_header(column)
        if normalized_key not in normalized_map:
            normalized_map[normalized_key] = column

    for candidate_name in candidate_names:
        normalized_candidate = normalize_consistency_header(candidate_name)
        if normalized_candidate in normalized_map:
            return normalized_map[normalized_candidate]
    return None

def normalize_consistency_dataframe(dataframe):
    if dataframe is None:
        return pd.DataFrame()

    output = dataframe.copy()
    if output.empty and len(output.columns) == 0:
        return output

    column_aliases = {
        CONS_COL_GROUP: ["URUN GRUBU"],
        CONS_COL_PAGE: ["SAYFA"],
        CONS_COL_DESC: ["URUN ACIKLAMASI"],
        CONS_COL_MATERIAL: ["MALZEME FIYATI"],
        CONS_COL_LABOR: ["ISCILIK FIYATI"],
        CONS_COL_GGK: ["GGK FIYATI"],
        CONS_COL_TOTAL: ["GENEL TOPLAM"],
        CONS_COL_COUNT: ["BULUNAN KAYIT SAYISI"],
        CONS_COL_QTY: ["MIKTAR", "TOPLAM MIKTAR"],
        CONS_COL_UNIT: ["BIRIM"],
        CONS_COL_MATERIAL_PCT: ["MALZEME %"],
        CONS_COL_LABOR_PCT: ["ISCILIK %"],
        CONS_COL_GGK_PCT: ["GGK %"],
        CONS_COL_TOTAL_PCT: ["GENEL TOPLAM %"],
        CONS_COL_UNIT_SCOPE: ["BIRIM FIYAT KAPSAMI"],
        CONS_COL_MATERIAL_UNIT_PRICE: ["MALZEME BIRIM FIYATI"],
        CONS_COL_LABOR_UNIT_PRICE: ["ISCILIK BIRIM FIYATI"],
        CONS_COL_GGK_UNIT_PRICE: ["GGK BIRIM FIYATI"],
        CONS_COL_TOTAL_UNIT_PRICE: ["GENEL TOPLAM BIRIM FIYATI"],
        CONS_COL_OFFER_CODE: ["TEKLIF KODU"],
        CONS_COL_OFFER_DATE: ["TEKLIF TARIHI"],
        CONS_COL_SELECTED_SHEETS: ["SECILEN SAYFALAR"],
        CONS_COL_SAVED_AT: ["KAYIT ZAMANI"],
    }

    rename_map = {}
    for canonical_name, candidates in column_aliases.items():
        resolved_name = resolve_consistency_column(output.columns, *candidates)
        if resolved_name and canonical_name not in output.columns:
            rename_map[resolved_name] = canonical_name

    return output.rename(columns=rename_map)

def coerce_memory_bool(series):
    normalized = series.fillna("").astype(str).str.strip().str.lower()
    return normalized.isin(["true", "1", "evet", "yes"])

def attach_variant_metadata(dataframe, variant_name):
    output = dataframe.copy()
    output[MEMORY_VARIANT_COL] = variant_name
    output[MEMORY_DEVICE_INCLUDED_COL] = variant_name == VARIANT_DEVICE_INCLUDED

    if "Teklif Kodu" in output.columns:
        offer_codes = output["Teklif Kodu"].fillna("").astype(str).str.strip().tolist()
        output[MEMORY_VARIANT_KEY_COL] = [
            build_variant_offer_key(offer_code, variant_name) for offer_code in offer_codes
        ]
    else:
        output[MEMORY_VARIANT_KEY_COL] = variant_name
    return output

def recompute_group_ratio_columns(group_df):
    output = group_df.copy()
    numeric_columns = [
        "Malzeme FiyatÄ±",
        "Ä°ÅŸÃ§ilik FiyatÄ±",
        "GGK FiyatÄ±",
        "Genel Toplam",
        "Bulunan KayÄ±t SayÄ±sÄ±",
    ]
    for column in numeric_columns:
        if column in output.columns:
            output[column] = pd.to_numeric(output[column], errors="coerce").fillna(0)

    percent_column_map = [
        ("Malzeme FiyatÄ±", "Malzeme %"),
        ("Ä°ÅŸÃ§ilik FiyatÄ±", "Ä°ÅŸÃ§ilik %"),
        ("GGK FiyatÄ±", "GGK %"),
        ("Genel Toplam", "Genel Toplam %"),
    ]
    for amount_column, percent_column in percent_column_map:
        if amount_column not in output.columns:
            continue
        total_value = float(pd.to_numeric(output[amount_column], errors="coerce").fillna(0).sum())
        if total_value > 0:
            output[percent_column] = (
                pd.to_numeric(output[amount_column], errors="coerce").fillna(0) / total_value * 100
            ).round(2)
        else:
            output[percent_column] = 0.0
    return output

def detect_offer_has_cihazlar(group_df):
    if group_df.empty or "ÃœrÃ¼n Grubu" not in group_df.columns or "Malzeme FiyatÄ±" not in group_df.columns:
        return False

    cihaz_mask = group_df["ÃœrÃ¼n Grubu"].apply(is_cihazlar_group)
    if not cihaz_mask.any():
        return False

    cihaz_malzeme = pd.to_numeric(group_df.loc[cihaz_mask, "Malzeme FiyatÄ±"], errors="coerce").fillna(0)
    return float(cihaz_malzeme.sum()) > 0

def build_group_snapshot_variants(group_df):
    offer_df = group_df.copy()
    has_cihazlar = detect_offer_has_cihazlar(offer_df)
    variants = {}

    if has_cihazlar:
        variants[VARIANT_DEVICE_INCLUDED] = attach_variant_metadata(
            recompute_group_ratio_columns(offer_df.copy()),
            VARIANT_DEVICE_INCLUDED,
        )
        excluded_df = offer_df[~offer_df["ÃœrÃ¼n Grubu"].apply(is_cihazlar_group)].copy()
    else:
        excluded_df = offer_df.copy()

    variants[VARIANT_DEVICE_EXCLUDED] = attach_variant_metadata(
        recompute_group_ratio_columns(excluded_df),
        VARIANT_DEVICE_EXCLUDED,
    )
    return variants, has_cihazlar

def build_item_snapshot_variants(item_df, has_cihazlar):
    offer_df = item_df.copy()
    variants = {}

    if has_cihazlar:
        variants[VARIANT_DEVICE_INCLUDED] = attach_variant_metadata(
            offer_df.copy(),
            VARIANT_DEVICE_INCLUDED,
        )
        excluded_df = offer_df[~offer_df["ÃœrÃ¼n Grubu"].apply(is_cihazlar_group)].copy()
    else:
        excluded_df = offer_df.copy()

    variants[VARIANT_DEVICE_EXCLUDED] = attach_variant_metadata(
        excluded_df,
        VARIANT_DEVICE_EXCLUDED,
    )
    return variants

def normalize_variant_memory(dataframe):
    if dataframe.empty:
        return dataframe.copy()

    output = dataframe.copy()
    if MEMORY_VARIANT_COL not in output.columns:
        output[MEMORY_VARIANT_COL] = VARIANT_DEVICE_EXCLUDED

    variant_series = output[MEMORY_VARIANT_COL].fillna("").astype(str).str.strip()
    variant_series = variant_series.replace("", VARIANT_DEVICE_EXCLUDED)
    output[MEMORY_VARIANT_COL] = variant_series.apply(
        lambda value: VARIANT_DEVICE_INCLUDED
        if "dahil" in turkce_ascii(value).lower()
        else VARIANT_DEVICE_EXCLUDED
    )

    if MEMORY_DEVICE_INCLUDED_COL in output.columns:
        output[MEMORY_DEVICE_INCLUDED_COL] = coerce_memory_bool(output[MEMORY_DEVICE_INCLUDED_COL])
    else:
        output[MEMORY_DEVICE_INCLUDED_COL] = output[MEMORY_VARIANT_COL] == VARIANT_DEVICE_INCLUDED

    if MEMORY_VARIANT_KEY_COL not in output.columns:
        if "Teklif Kodu" in output.columns:
            offer_codes = output["Teklif Kodu"].fillna("").astype(str).str.strip().tolist()
            output[MEMORY_VARIANT_KEY_COL] = [
                build_variant_offer_key(offer_code, variant_name)
                for offer_code, variant_name in zip(offer_codes, output[MEMORY_VARIANT_COL].tolist())
            ]
        else:
            output[MEMORY_VARIANT_KEY_COL] = output[MEMORY_VARIANT_COL]
    return output

def expand_group_memory_variants(memory_df):
    if memory_df.empty:
        return memory_df.copy(), {}

    if MEMORY_VARIANT_COL in memory_df.columns or MEMORY_VARIANT_KEY_COL in memory_df.columns:
        normalized = normalize_variant_memory(memory_df)
        if "Teklif Kodu" in normalized.columns:
            device_status_by_offer = (
                normalized.groupby("Teklif Kodu")[MEMORY_DEVICE_INCLUDED_COL].max().astype(bool).to_dict()
            )
        else:
            device_status_by_offer = {}
        return normalized, {str(key).strip(): bool(value) for key, value in device_status_by_offer.items()}

    if "Teklif Kodu" not in memory_df.columns:
        return normalize_variant_memory(memory_df), {}

    expanded_frames = []
    device_status_by_offer = {}
    for offer_code, offer_df in memory_df.groupby("Teklif Kodu", sort=False):
        variants, has_cihazlar = build_group_snapshot_variants(offer_df.copy())
        device_status_by_offer[str(offer_code).strip()] = has_cihazlar
        expanded_frames.extend(list(variants.values()))

    if not expanded_frames:
        return normalize_variant_memory(memory_df), device_status_by_offer
    return pd.concat(expanded_frames, ignore_index=True), device_status_by_offer

def expand_item_memory_variants(memory_df, device_status_by_offer):
    if memory_df.empty:
        return memory_df.copy()

    if MEMORY_VARIANT_COL in memory_df.columns or MEMORY_VARIANT_KEY_COL in memory_df.columns:
        return normalize_variant_memory(memory_df)

    if "Teklif Kodu" not in memory_df.columns:
        return normalize_variant_memory(memory_df)

    expanded_frames = []
    for offer_code, offer_df in memory_df.groupby("Teklif Kodu", sort=False):
        has_cihazlar = bool(device_status_by_offer.get(str(offer_code).strip(), False))
        variants = build_item_snapshot_variants(offer_df.copy(), has_cihazlar)
        expanded_frames.extend(list(variants.values()))

    if not expanded_frames:
        return normalize_variant_memory(memory_df)
    return pd.concat(expanded_frames, ignore_index=True)

def sum_numeric_column(dataframe, column_name):
    if dataframe is None or dataframe.empty or column_name not in dataframe.columns:
        return 0.0
    return float(pd.to_numeric(dataframe[column_name], errors="coerce").fillna(0).sum())

def render_consistency_module_legacy(all_df, summary_data, selected_sheets, source_file_name):
    render_section_heading("Tutarlılık ve Geçmiş Kıyas Modülü", icon="")
    st.markdown(
        '<div class="info-card">Bu modül teklif verilerini hafızaya kaydeder, geçmiş tekliflerle kıyaslar ve olası eksik maliyet girişlerini anomali olarak işaretler.</div>',
        unsafe_allow_html=True,
    )

    teklif_kodu = derive_offer_code_from_outputs(
        all_df,
        selected_sheets=selected_sheets,
        fallback_name=source_file_name,
    )
    teklif_tarihi = pd.Timestamp.now().normalize()

    group_memory_path, item_memory_path, system_memory_path = get_offer_memory_paths()
    history_groups = load_memory_dataframe(group_memory_path)
    history_items = load_memory_dataframe(item_memory_path)
    history_systems = load_memory_dataframe(system_memory_path)

    context_groups = history_groups.copy()
    context_items = history_items.copy()
    context_systems = history_systems.copy()
    context_groups_main_category = context_groups.copy()
    context_items_main_category = context_items.copy()
    context_systems_main_category = context_systems.copy()

    st.caption(
        f"Hafızadaki kayıtlar: Grup {len(context_groups)} satır, "
        f"Ürün {len(context_items)} satır, Sistem {len(context_systems)} satır"
    )

    current_group_snapshot = summary_data[
        ["Ürün Grubu", "Malzeme Fiyatı", "İşçilik Fiyatı", "GGK Fiyatı", "Genel Toplam", "Bulunan Kayıt Sayısı", "Malzeme %", "İşçilik %", "GGK %", "Genel Toplam %"]
    ].copy()
    current_group_snapshot["Teklif Kodu"] = teklif_kodu
    current_group_snapshot["Teklif Tarihi"] = pd.to_datetime(teklif_tarihi)
    current_group_snapshot["Seçilen Sayfalar"] = " | ".join(selected_sheets)
    current_group_snapshot["Kayıt Zamanı"] = pd.Timestamp.now()

    current_item_snapshot = all_df[
        ["Ürün Grubu", "Sayfa", "Ürün Açıklaması", "Malzeme Fiyatı", "İşçilik Fiyatı", "GGK Fiyatı", "Genel Toplam"]
    ].copy()
    current_item_snapshot = current_item_snapshot.groupby(
        ["Ürün Grubu", "Sayfa", "Ürün Açıklaması"], as_index=False
    ).agg(
        {
            "Malzeme Fiyatı": "sum",
            "İşçilik Fiyatı": "sum",
            "GGK Fiyatı": "sum",
            "Genel Toplam": "sum",
        }
    )
    current_item_snapshot["Teklif Kodu"] = teklif_kodu
    current_item_snapshot["Teklif Tarihi"] = pd.to_datetime(teklif_tarihi)
    current_item_snapshot["Kayıt Zamanı"] = pd.Timestamp.now()

    st.markdown("### Mevcut Teklif Yüzdelik Dağılım")
    toplam_malzeme = pd.to_numeric(all_df["Malzeme Fiyatı"], errors="coerce").fillna(0).sum()
    toplam_iscilik = pd.to_numeric(all_df["İşçilik Fiyatı"], errors="coerce").fillna(0).sum()
    toplam_ggk = pd.to_numeric(all_df["GGK Fiyatı"], errors="coerce").fillna(0).sum()
    toplam_genel = pd.to_numeric(all_df["Genel Toplam"], errors="coerce").fillna(0).sum()
    mevcut_dagilim_df = pd.DataFrame(
        [
            {"Kalem": "Malzeme", "Tutar": toplam_malzeme},
            {"Kalem": "İşçilik", "Tutar": toplam_iscilik},
            {"Kalem": "GGK", "Tutar": toplam_ggk},
        ]
    )
    if toplam_genel > 0:
        mevcut_dagilim_df["Teklif İçindeki %"] = (mevcut_dagilim_df["Tutar"] / toplam_genel * 100).round(2)
    else:
        mevcut_dagilim_df["Teklif İçindeki %"] = 0.0
    st.dataframe(mevcut_dagilim_df, use_container_width=True, hide_index=True, height=180)

    st.markdown("### Tesisat/Sistem Dağılımı ve Medyan Kıyas")
    current_system_snapshot = build_system_distribution(all_df)
    selected_system_labels = []
    for category in SYSTEM_TYPE_DEFINITIONS.keys():
        category_options = current_system_snapshot[
            current_system_snapshot["Sistem Kategori"] == category
        ]["Sistem Tipi"].tolist()
        selector_key = f"included_systems_{normalize_text_for_match(category).replace(' ', '_')}"
        selected_types = st.multiselect(
            category,
            options=category_options,
            default=category_options,
            key=selector_key,
        )
        for selected_type in selected_types:
            selected_system_labels.append(f"{category} / {selected_type}")

    if not selected_system_labels:
        st.warning("En az bir sistem tipi seçilmediği için tüm sistem tipleri üzerinden devam ediliyor.")
        selected_system_labels = current_system_snapshot["Sistem Etiket"].tolist()

    selected_system_set = set(selected_system_labels)
    current_system_view = current_system_snapshot[
        current_system_snapshot["Sistem Etiket"].isin(selected_system_labels)
    ].copy()
    selected_system_total = current_system_view["Sistem Toplam"].sum()
    if selected_system_total > 0:
        current_system_view["Seçilen Sistem Seti %"] = (
            current_system_view["Sistem Toplam"] / selected_system_total * 100
        ).round(2)
    else:
        current_system_view["Seçilen Sistem Seti %"] = 0.0

    st.dataframe(
        current_system_view[
            ["Sistem Kategori", "Sistem Tipi", "Sistem Toplam", "Sistem %", "Seçilen Sistem Seti %"]
        ].sort_values("Sistem Toplam", ascending=False),
        use_container_width=True,
        hide_index=True,
        height=320,
    )

    sistem_eslesen_toplam = current_system_snapshot["Sistem Toplam"].sum()
    sistem_eslesme_yuzde = (sistem_eslesen_toplam / toplam_genel * 100) if toplam_genel > 0 else 0.0
    st.caption(
        f"Sistem eşleşme kapsaması: {format_currency_display(sistem_eslesen_toplam)} "
        f"({sistem_eslesme_yuzde:.2f}%) / Toplam {format_currency_display(toplam_genel)}"
    )
    if toplam_genel > 0 and sistem_eslesen_toplam <= 0:
        st.info("Sistem dağılımı anahtar kelime eşleşmesi bulamadı. Ürün açıklamalarında sistem anahtarları geçmiyor olabilir.")

    current_system_snapshot["Teklife Dahil"] = current_system_snapshot["Sistem Etiket"].isin(selected_system_set)
    current_system_snapshot["Sistem"] = current_system_snapshot["Sistem Etiket"]
    current_system_snapshot["Teklif Kodu"] = teklif_kodu
    current_system_snapshot["Teklif Tarihi"] = pd.to_datetime(teklif_tarihi)
    current_system_snapshot["Seçilen Sayfalar"] = " | ".join(selected_sheets)
    current_system_snapshot["Kayıt Zamanı"] = pd.Timestamp.now()

    system_ratio_threshold = st.slider(
        "Sistem dağılımı düşük oran eşiği (%)",
        min_value=5,
        max_value=60,
        value=20,
        step=5,
        key="system_ratio_threshold",
    )

    st.caption("Sistem medyan kıyası, hafızadaki tüm geçmiş teklifler üzerinden yapılır.")
    if not context_systems_main_category.empty:
        reference_systems = context_systems_main_category[context_systems_main_category["Teklif Kodu"] != teklif_kodu].copy()
    else:
        reference_systems = pd.DataFrame()

    if not reference_systems.empty:
        if "Teklife Dahil" in reference_systems.columns:
            included_mask = (
                reference_systems["Teklife Dahil"]
                .astype(str)
                .str.strip()
                .str.lower()
                .isin(["true", "1", "evet", "yes"])
            )
            reference_systems = reference_systems[included_mask]

        if "Sistem Etiket" not in reference_systems.columns:
            if "Sistem" in reference_systems.columns:
                reference_systems["Sistem Etiket"] = reference_systems["Sistem"].astype(str)
            elif "Sistem Kategori" in reference_systems.columns and "Sistem Tipi" in reference_systems.columns:
                reference_systems["Sistem Etiket"] = (
                    reference_systems["Sistem Kategori"].astype(str) + " / " + reference_systems["Sistem Tipi"].astype(str)
                )
            else:
                reference_systems["Sistem Etiket"] = ""

        reference_systems = reference_systems[reference_systems["Sistem Etiket"].isin(selected_system_labels)].copy()
        if reference_systems.empty:
            st.info("Seçilen sistemler için bu bağlamda referans teklif verisi henüz yok.")
        else:
            system_medians = reference_systems.groupby("Sistem Etiket")["Sistem %"].median().reset_index(name="Sistem % Medyan")
            system_reference_counts = reference_systems.groupby("Sistem Etiket")["Teklif Kodu"].nunique().reset_index(name="Referans Teklif Sayısı")

            system_compare = current_system_view[
                ["Sistem Kategori", "Sistem Tipi", "Sistem Etiket", "Sistem %"]
            ].rename(columns={"Sistem %": "Sistem % Mevcut"})
            system_compare = system_compare.merge(system_medians, on="Sistem Etiket", how="left")
            system_compare = system_compare.merge(system_reference_counts, on="Sistem Etiket", how="left")
            system_compare["Referans Teklif Sayısı"] = system_compare["Referans Teklif Sayısı"].fillna(0).astype(int)

            global_system_median = reference_systems["Sistem %"].median()
            system_compare["Sistem % Medyan"] = system_compare["Sistem % Medyan"].fillna(global_system_median).round(2)
            system_compare["Sistem % Sapma"] = (system_compare["Sistem % Mevcut"] - system_compare["Sistem % Medyan"]).round(2)
            system_compare["Düşük Oran"] = system_compare["Sistem % Mevcut"] < (
                system_compare["Sistem % Medyan"] * (1 - system_ratio_threshold / 100)
            )
            system_compare = system_compare.sort_values(["Düşük Oran", "Sistem % Sapma"], ascending=[False, True])

            st.markdown("#### Sistem Dağılımı: Mevcut vs Geçmiş Medyan")
            st.dataframe(
                system_compare[
                    ["Sistem Kategori", "Sistem Tipi", "Sistem % Mevcut", "Sistem % Medyan", "Sistem % Sapma", "Referans Teklif Sayısı", "Düşük Oran"]
                ],
                use_container_width=True,
                hide_index=True,
                height=320
            )

            low_systems = system_compare[system_compare["Düşük Oran"]].copy()
            if low_systems.empty:
                st.success("Seçilen tesisat/sistem dağılımında kritik düşük oran tespit edilmedi.")
            else:
                st.warning(f"{len(low_systems)} sistemde medyana göre düşük oran anomalisi tespit edildi.")
                st.dataframe(low_systems, use_container_width=True, hide_index=True, height=260)
    else:
        st.info("Sistem dağılımı medyan kıyası için bu bağlamda hafıza verisi henüz yok.")

    st.markdown("### Teklifler Arası Ürün Grubu % Kıyas")
    st.caption("Ürün grubu % kıyası, hafızadaki tüm geçmiş teklifler üzerinden yapılır.")
    if not context_groups_main_category.empty:
        historical_groups = context_groups_main_category[context_groups_main_category["Teklif Kodu"] != teklif_kodu].copy()
    else:
        historical_groups = pd.DataFrame()

    if historical_groups.empty:
        st.info("Kıyaslama için geçmiş teklif verisi henüz yok. Önce teklifleri hafızaya kaydedin.")
    else:
        ratio_cols = ["Genel Toplam %", "Malzeme %", "İşçilik %", "GGK %"]
        offer_dates = historical_groups.groupby("Teklif Kodu", as_index=False)["Teklif Tarihi"].max()
        historical_offer_table = historical_groups.pivot_table(
            index="Teklif Kodu",
            columns="Ürün Grubu",
            values="Genel Toplam %",
            aggfunc="sum",
            fill_value=0.0
        ).reset_index()
        historical_offer_table = historical_offer_table.merge(offer_dates, on="Teklif Kodu", how="left")
        ordered_cols = ["Teklif Kodu", "Teklif Tarihi"] + [
            col for col in historical_offer_table.columns if col not in ["Teklif Kodu", "Teklif Tarihi"]
        ]

        st.markdown("#### Geçmiş Tekliflerin Ürün Grubu Yüzdeleri")
        st.dataframe(
            historical_offer_table[ordered_cols].sort_values("Teklif Tarihi", ascending=False),
            use_container_width=True,
            hide_index=True,
            height=280
        )

        compare_mode = st.radio(
            "Kıyas Kaynağı",
            options=["Tüm geçmiş teklifler ortalaması", "Seçilen teklifler"],
            horizontal=True,
            key="offer_compare_mode",
        )
        if compare_mode == "Seçilen teklifler":
            offer_options = historical_offer_table["Teklif Kodu"].astype(str).tolist()
            selected_offers = st.multiselect(
                "Kıyas için teklif seçin",
                options=offer_options,
                default=offer_options[:1] if offer_options else [],
                key="selected_offer_codes",
            )
            benchmark_groups = historical_groups[historical_groups["Teklif Kodu"].astype(str).isin(selected_offers)].copy()
        else:
            benchmark_groups = historical_groups.copy()

        if benchmark_groups.empty:
            st.warning("Seçilen kıyas kaynağında veri yok.")
        else:
            benchmark_avg = benchmark_groups.groupby("Ürün Grubu")[ratio_cols].mean().reset_index()
            benchmark_count = benchmark_groups.groupby("Ürün Grubu")["Teklif Kodu"].nunique().reset_index(name="Referans Teklif Sayısı")

            comparison_table = current_group_snapshot[["Ürün Grubu"] + ratio_cols].merge(
                benchmark_avg,
                on="Ürün Grubu",
                how="left",
                suffixes=(" Mevcut", " Ortalama")
            )
            comparison_table = comparison_table.merge(benchmark_count, on="Ürün Grubu", how="left")
            comparison_table["Referans Teklif Sayısı"] = comparison_table["Referans Teklif Sayısı"].fillna(0).astype(int)

            global_avg = benchmark_groups[ratio_cols].mean()
            for ratio_col in ratio_cols:
                comparison_table[f"{ratio_col} Ortalama"] = comparison_table[f"{ratio_col} Ortalama"].fillna(global_avg.get(ratio_col, np.nan))
                comparison_table[f"{ratio_col} Fark"] = (
                    comparison_table[f"{ratio_col} Mevcut"] - comparison_table[f"{ratio_col} Ortalama"]
                ).round(2)

            deviation_threshold = st.slider(
                "Genel Toplam % sapma eşiği (puan)",
                min_value=1,
                max_value=50,
                value=10,
                step=1,
                key="offer_ratio_deviation_threshold",
            )
            comparison_table["Yüksek Sapma"] = comparison_table["Genel Toplam % Fark"].abs() >= deviation_threshold
            comparison_table = comparison_table.sort_values("Genel Toplam % Fark", key=lambda s: s.abs(), ascending=False)

            st.markdown("#### Mevcut Teklif vs Kıyas Kaynağı (Ortalama)")
            st.markdown(
                """
                <div style="display:flex; gap:12px; flex-wrap:wrap; margin-bottom:8px;">
                  <span style="background:#dbeafe; color:#1e3a8a; padding:4px 10px; border-radius:999px; font-weight:700;">Mevcut %</span>
                  <span style="background:#dcfce7; color:#166534; padding:4px 10px; border-radius:999px; font-weight:700;">Kıyas Ortalama %</span>
                  <span style="background:#ffedd5; color:#9a3412; padding:4px 10px; border-radius:999px; font-weight:700;">Fark % Puan</span>
                  <span style="background:#fee2e2; color:#991b1b; padding:4px 10px; border-radius:999px; font-weight:700;">Yüksek Sapma</span>
                </div>
                """,
                unsafe_allow_html=True
            )

            comparison_display = comparison_table.copy()
            comparison_display["Sapma Durumu"] = np.where(comparison_display["Yüksek Sapma"], "Yüksek Sapma", "Normal")

            ordered_cols = ["Ürün Grubu"]
            for ratio_col in ratio_cols:
                ordered_cols.extend([f"{ratio_col} Mevcut", f"{ratio_col} Ortalama", f"{ratio_col} Fark"])
            ordered_cols.extend(["Referans Teklif Sayısı", "Sapma Durumu", "Yüksek Sapma"])
            comparison_display = comparison_display[ordered_cols]

            percent_cols = [col for col in comparison_display.columns if "%" in col]
            formatters = {}
            for col in percent_cols:
                if col.endswith(" Fark"):
                    formatters[col] = lambda v: f"{float(v):+.2f} puan" if pd.notna(v) else ""
                else:
                    formatters[col] = lambda v: f"{float(v):.2f}%" if pd.notna(v) else ""

            def style_offer_comparison(df):
                styles = pd.DataFrame("", index=df.index, columns=df.columns)
                for col in df.columns:
                    if col.endswith(" Mevcut"):
                        styles[col] = "background-color: #dbeafe; color: #1e3a8a; font-weight: 700; text-align:center;"
                    elif col.endswith(" Ortalama"):
                        styles[col] = "background-color: #dcfce7; color: #166534; font-weight: 700; text-align:center;"
                    elif col.endswith(" Fark"):
                        styles[col] = df[col].apply(
                            lambda x: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                            if pd.notna(x) and float(x) >= 0
                            else "background-color: #dcfce7; color: #166534; font-weight: 800; text-align:center;"
                        )
                    elif col == "Sapma Durumu":
                        styles[col] = df["Yüksek Sapma"].apply(
                            lambda is_high: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                            if bool(is_high)
                            else "background-color: #f8fafc; color: #334155; font-weight: 700; text-align:center;"
                        )
                    elif "%" in col:
                        styles[col] = "text-align:center; font-weight:700;"
                return styles

            styled_comparison = (
                comparison_display.style
                .apply(style_offer_comparison, axis=None)
                .format(formatters)
            )
            st.dataframe(
                styled_comparison,
                use_container_width=True,
                hide_index=True,
                height=420
            )

            high_deviation_rows = comparison_table[comparison_table["Yüksek Sapma"]]
            if high_deviation_rows.empty:
                st.success("Genel Toplam % açısından yüksek sapma tespit edilmedi.")
            else:
                st.warning(f"{len(high_deviation_rows)} ürün grubunda yüksek sapma tespit edildi.")

    st.markdown("### Geçmiş Dönem Kıyaslaması (Aynı Ürün/Sayfa)")
    st.caption("Bu kıyas, hafızadaki tüm geçmiş teklifler üzerinden yapılır.")
    drop_threshold = st.slider("Sert düşüş eşiği (%)", min_value=5, max_value=80, value=20, step=5, key="drop_threshold")

    if not context_items_main_category.empty:
        reference_items = context_items_main_category[context_items_main_category["Teklif Kodu"] != teklif_kodu].copy()
    else:
        reference_items = pd.DataFrame()

    if not reference_items.empty:
        key_cols = ["Ürün Grubu", "Sayfa", "Ürün Açıklaması"]
        reference_items = reference_items.sort_values("Teklif Tarihi")
        latest_reference = reference_items.groupby(key_cols, as_index=False).tail(1)
        latest_reference = latest_reference[key_cols + ["Genel Toplam", "Teklif Kodu", "Teklif Tarihi"]].rename(
            columns={
                "Genel Toplam": "Önceki Genel Toplam",
                "Teklif Kodu": "Önceki Teklif Kodu",
                "Teklif Tarihi": "Önceki Teklif Tarihi",
            }
        )

        comparison_df = current_item_snapshot[key_cols + ["Genel Toplam"]].rename(columns={"Genel Toplam": "Mevcut Genel Toplam"})
        comparison_df = comparison_df.merge(latest_reference, on=key_cols, how="left")
        comparison_df = comparison_df[comparison_df["Önceki Genel Toplam"].notna()].copy()

        if comparison_df.empty:
            st.info("Aynı ürün/sayfa için kıyaslanacak geçmiş kayıt bulunamadı.")
        else:
            comparison_df["Değişim %"] = ((comparison_df["Mevcut Genel Toplam"] - comparison_df["Önceki Genel Toplam"]) / comparison_df["Önceki Genel Toplam"] * 100).round(2)
            comparison_df["Sert Düşüş"] = comparison_df["Değişim %"] <= (-drop_threshold)
            sharp_drops = comparison_df[comparison_df["Sert Düşüş"]].sort_values("Değişim %")

            if sharp_drops.empty:
                st.success("Sert düşüş eşiğini aşan anomali tespit edilmedi.")
            else:
                st.warning(f"{len(sharp_drops)} kalemde geçmiş döneme göre sert düşüş tespit edildi.")
                st.dataframe(sharp_drops, use_container_width=True, hide_index=True, height=360)
    else:
        st.info("Geçmiş dönem kıyaslaması için bu bağlamda hafıza verisi henüz yok.")

    st.markdown("### Teklif Hafızasına Kaydet")
    st.caption("Kaydedilen kayıtlar, ileride otomatik kıyaslamalarda kullanılacaktır.")

    if st.button("Bu Teklifi Hafızaya Kaydet", key="save_offer_memory"):
        if not teklif_kodu:
            st.error("Lütfen geçerli bir Teklif Kodu girin.")
        else:
            try:
                upsert_memory_dataframe(
                    group_memory_path,
                    current_group_snapshot,
                    key_columns=["Teklif Kodu", "Ürün Grubu"],
                )
                upsert_memory_dataframe(
                    item_memory_path,
                    current_item_snapshot,
                    key_columns=["Teklif Kodu", "Ürün Grubu", "Sayfa", "Ürün Açıklaması"],
                )
                upsert_memory_dataframe(
                    system_memory_path,
                    current_system_snapshot,
                    key_columns=["Teklif Kodu", "Sistem Etiket"],
                )
                st.success("Teklif hafızaya kaydedildi ve mevcut kayıtlar güncellendi.")
            except Exception as save_error:
                st.error(f"Hafıza kaydı sırasında hata: {str(save_error)}")

def render_consistency_module(all_df, summary_data, selected_sheets, source_file_name):
    render_section_heading("Tutarlılık ve Geçmiş Kıyas Modülü", icon="")
    st.markdown(
        '<div class="info-card">Bu modül teklif verilerini hafızaya kaydeder, geçmiş tekliflerle kıyaslar ve CİHAZLAR dahil/hariç durumuna göre ayrı referans havuzları kullanır.</div>',
        unsafe_allow_html=True,
    )

    teklif_kodu = derive_offer_code_from_outputs(
        all_df,
        selected_sheets=selected_sheets,
        fallback_name=source_file_name,
    )
    teklif_tarihi = pd.Timestamp.now().normalize()
    kayit_zamani = pd.Timestamp.now()

    render_consistency_hero(
        source_file_name=source_file_name,
        teklif_kodu=teklif_kodu,
        selected_sheets=selected_sheets,
        current_offer_has_cihazlar=detect_offer_has_cihazlar(normalized_summary),
    )

    group_memory_path, item_memory_path, _ = get_offer_memory_paths()
    history_groups_raw = load_memory_dataframe(group_memory_path)
    history_items_raw = load_memory_dataframe(item_memory_path)
    history_groups, device_status_by_offer = expand_group_memory_variants(history_groups_raw)
    history_items = expand_item_memory_variants(history_items_raw, device_status_by_offer)

    current_group_base = summary_data[
        ["Ürün Grubu", "Malzeme Fiyatı", "İşçilik Fiyatı", "GGK Fiyatı", "Genel Toplam", "Bulunan Kayıt Sayısı", "Malzeme %", "İşçilik %", "GGK %", "Genel Toplam %"]
    ].copy()
    current_group_base["Teklif Kodu"] = teklif_kodu
    current_group_base["Teklif Tarihi"] = pd.to_datetime(teklif_tarihi)
    current_group_base["Seçilen Sayfalar"] = " | ".join(selected_sheets)
    current_group_base["Kayıt Zamanı"] = kayit_zamani

    current_item_base = all_df[
        ["Ürün Grubu", "Sayfa", "Ürün Açıklaması", "Malzeme Fiyatı", "İşçilik Fiyatı", "GGK Fiyatı", "Genel Toplam"]
    ].copy()
    current_item_base = current_item_base.groupby(
        ["Ürün Grubu", "Sayfa", "Ürün Açıklaması"], as_index=False
    ).agg(
        {
            "Malzeme Fiyatı": "sum",
            "İşçilik Fiyatı": "sum",
            "GGK Fiyatı": "sum",
            "Genel Toplam": "sum",
        }
    )
    current_item_base["Teklif Kodu"] = teklif_kodu
    current_item_base["Teklif Tarihi"] = pd.to_datetime(teklif_tarihi)
    current_item_base["Kayıt Zamanı"] = kayit_zamani

    current_group_variants, current_offer_has_cihazlar = build_group_snapshot_variants(current_group_base)
    current_item_variants = build_item_snapshot_variants(current_item_base, current_offer_has_cihazlar)

    cihaz_mask = current_group_base["Ürün Grubu"].apply(is_cihazlar_group)
    cihaz_malzeme_toplami = sum_numeric_column(current_group_base.loc[cihaz_mask].copy(), "Malzeme Fiyatı")

    analysis_variant_options = [VARIANT_DEVICE_EXCLUDED]
    if current_offer_has_cihazlar:
        analysis_variant_options = [VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED]

    selected_analysis_variant = st.radio(
        "CİHAZLAR analiz modu",
        options=analysis_variant_options,
        horizontal=True,
        key="cihazlar_analysis_variant",
    )

    current_group_snapshot = current_group_variants[selected_analysis_variant].copy()
    current_item_snapshot = current_item_variants[selected_analysis_variant].copy()
    analysis_offer_key = build_variant_offer_key(teklif_kodu, selected_analysis_variant)

    context_groups = history_groups.copy()
    context_items = history_items.copy()
    if not context_groups.empty:
        context_groups = context_groups[context_groups[MEMORY_VARIANT_COL] == selected_analysis_variant].copy()
    if not context_items.empty:
        context_items = context_items[context_items[MEMORY_VARIANT_COL] == selected_analysis_variant].copy()

    st.caption(
        f"CİHAZLAR malzeme toplamı: {format_currency_display(cihaz_malzeme_toplami)} | "
        f"Aktif kıyas modu: {selected_analysis_variant} | "
        f"Hafıza satırları: Grup {len(context_groups)}, Ürün {len(context_items)}"
    )
    if current_offer_has_cihazlar:
        st.info("Bu teklifte CİHAZLAR malzemesi bulundu. Kıyas ve hafıza kayıtları cihaz dahil ve cihaz hariç olarak ayrı yürür.")
    else:
        st.info("Bu teklifte CİHAZLAR malzeme toplamı 0 olduğu için cihaz hariç referans havuzu kullanılacaktır.")

    st.markdown("### Mevcut Teklif Yüzdelik Dağılım")
    toplam_malzeme = sum_numeric_column(current_item_snapshot, "Malzeme Fiyatı")
    toplam_iscilik = sum_numeric_column(current_item_snapshot, "İşçilik Fiyatı")
    toplam_ggk = sum_numeric_column(current_item_snapshot, "GGK Fiyatı")
    toplam_genel = sum_numeric_column(current_item_snapshot, "Genel Toplam")
    mevcut_dagilim_df = pd.DataFrame(
        [
            {"Kalem": "Malzeme", "Tutar": toplam_malzeme},
            {"Kalem": "İşçilik", "Tutar": toplam_iscilik},
            {"Kalem": "GGK", "Tutar": toplam_ggk},
        ]
    )
    if toplam_genel > 0:
        mevcut_dagilim_df["Teklif İçindeki %"] = (mevcut_dagilim_df["Tutar"] / toplam_genel * 100).round(2)
    else:
        mevcut_dagilim_df["Teklif İçindeki %"] = 0.0
    create_sortable_numeric_table(
        mevcut_dagilim_df,
        height=180,
        currency_cols=["Tutar"],
        percent_cols=["Teklif İçindeki %"],
    )

    render_consistency_section_banner(
        "Teklifler Arasi Urun Grubu Yuzde Kiyasi",
        "Mevcut teklif dagilimini gecmis ortalamalarla yan yana okuyun ve sapma olan alanlari hizli sekilde tespit edin.",
        tone="blue",
        eyebrow="Gecmis Ortalama",
    )
    st.caption(f"Ürün grubu % kıyası, hafızadaki {selected_analysis_variant.lower()} varyantı üzerinden yapılır.")

    historical_groups = context_groups.copy()
    if not historical_groups.empty:
        historical_groups = historical_groups[historical_groups[MEMORY_VARIANT_KEY_COL] != analysis_offer_key].copy()

    if historical_groups.empty:
        st.info("Kıyaslama için bu cihaz modunda geçmiş teklif verisi henüz yok. Önce teklifleri hafızaya kaydedin.")
    else:
        ratio_cols = ["Genel Toplam %", "Malzeme %", "İşçilik %", "GGK %"]
        offer_dates = historical_groups.groupby(MEMORY_VARIANT_KEY_COL, as_index=False)["Teklif Tarihi"].max()
        offer_meta = historical_groups.groupby(MEMORY_VARIANT_KEY_COL, as_index=False).agg(
            {
                "Teklif Kodu": "first",
                MEMORY_VARIANT_COL: "first",
            }
        )
        historical_offer_table = historical_groups.pivot_table(
            index=MEMORY_VARIANT_KEY_COL,
            columns="Ürün Grubu",
            values="Genel Toplam %",
            aggfunc="sum",
            fill_value=0.0
        ).reset_index()
        historical_offer_table = historical_offer_table.merge(offer_meta, on=MEMORY_VARIANT_KEY_COL, how="left")
        historical_offer_table = historical_offer_table.merge(offer_dates, on=MEMORY_VARIANT_KEY_COL, how="left")

        ordered_cols = [MEMORY_VARIANT_KEY_COL, "Teklif Kodu", MEMORY_VARIANT_COL, "Teklif Tarihi"] + [
            col for col in historical_offer_table.columns
            if col not in [MEMORY_VARIANT_KEY_COL, "Teklif Kodu", MEMORY_VARIANT_COL, "Teklif Tarihi"]
        ]

        st.markdown("#### Geçmiş Tekliflerin Ürün Grubu Yüzdeleri")
        st.dataframe(
            historical_offer_table[ordered_cols].sort_values("Teklif Tarihi", ascending=False),
            use_container_width=True,
            hide_index=True,
            height=280
        )

        compare_mode = st.radio(
            "Kıyas Kaynağı",
            options=["Tüm geçmiş teklifler ortalaması", "Seçilen teklifler"],
            horizontal=True,
            key="offer_compare_mode",
        )
        if compare_mode == "Seçilen teklifler":
            offer_options = historical_offer_table[MEMORY_VARIANT_KEY_COL].astype(str).tolist()
            selected_offers = st.multiselect(
                "Kıyas için teklif seçin",
                options=offer_options,
                default=offer_options[:1] if offer_options else [],
                key="selected_offer_codes",
            )
            benchmark_groups = historical_groups[
                historical_groups[MEMORY_VARIANT_KEY_COL].astype(str).isin(selected_offers)
            ].copy()
        else:
            benchmark_groups = historical_groups.copy()

        if benchmark_groups.empty:
            st.warning("Seçilen kıyas kaynağında veri yok.")
        elif current_group_snapshot.empty:
            st.info("Seçilen cihaz modunda kıyaslanacak ürün grubu kalmadı.")
        else:
            benchmark_avg = benchmark_groups.groupby("Ürün Grubu")[ratio_cols].mean().reset_index()
            benchmark_count = benchmark_groups.groupby("Ürün Grubu")[MEMORY_VARIANT_KEY_COL].nunique().reset_index(name="Referans Teklif Sayısı")

            comparison_table = current_group_snapshot[["Ürün Grubu"] + ratio_cols].merge(
                benchmark_avg,
                on="Ürün Grubu",
                how="left",
                suffixes=(" Mevcut", " Ortalama")
            )
            comparison_table = comparison_table.merge(benchmark_count, on="Ürün Grubu", how="left")
            comparison_table["Referans Teklif Sayısı"] = comparison_table["Referans Teklif Sayısı"].fillna(0).astype(int)

            global_avg = benchmark_groups[ratio_cols].mean()
            for ratio_col in ratio_cols:
                comparison_table[f"{ratio_col} Ortalama"] = comparison_table[f"{ratio_col} Ortalama"].fillna(global_avg.get(ratio_col, np.nan))
                comparison_table[f"{ratio_col} Fark"] = (
                    comparison_table[f"{ratio_col} Mevcut"] - comparison_table[f"{ratio_col} Ortalama"]
                ).round(2)

            deviation_threshold = st.slider(
                "Genel Toplam % sapma eşiği (puan)",
                min_value=1,
                max_value=50,
                value=10,
                step=1,
                key="offer_ratio_deviation_threshold",
            )
            comparison_table["Yüksek Sapma"] = comparison_table["Genel Toplam % Fark"].abs() >= deviation_threshold
            comparison_table = comparison_table.sort_values("Genel Toplam % Fark", key=lambda s: s.abs(), ascending=False)

            st.markdown("#### Mevcut Teklif vs Kıyas Kaynağı (Ortalama)")
            st.markdown(
                """
                <div style="display:flex; gap:12px; flex-wrap:wrap; margin-bottom:8px;">
                  <span style="background:#dbeafe; color:#1e3a8a; padding:4px 10px; border-radius:999px; font-weight:700;">Mevcut %</span>
                  <span style="background:#dcfce7; color:#166534; padding:4px 10px; border-radius:999px; font-weight:700;">Kıyas Ortalama %</span>
                  <span style="background:#ffedd5; color:#9a3412; padding:4px 10px; border-radius:999px; font-weight:700;">Fark % Puan</span>
                  <span style="background:#fee2e2; color:#991b1b; padding:4px 10px; border-radius:999px; font-weight:700;">Yüksek Sapma</span>
                </div>
                """,
                unsafe_allow_html=True
            )

            comparison_display = comparison_table.copy()
            comparison_display["Sapma Durumu"] = np.where(comparison_display["Yüksek Sapma"], "Yüksek Sapma", "Normal")

            ordered_cols = ["Ürün Grubu"]
            for ratio_col in ratio_cols:
                ordered_cols.extend([f"{ratio_col} Mevcut", f"{ratio_col} Ortalama", f"{ratio_col} Fark"])
            ordered_cols.extend(["Referans Teklif Sayısı", "Sapma Durumu", "Yüksek Sapma"])
            comparison_display = comparison_display[ordered_cols]

            percent_cols = [col for col in comparison_display.columns if "%" in col]
            formatters = {}
            for col in percent_cols:
                if col.endswith(" Fark"):
                    formatters[col] = lambda v: f"{float(v):+.2f} puan" if pd.notna(v) else ""
                else:
                    formatters[col] = lambda v: f"{float(v):.2f}%" if pd.notna(v) else ""

            def style_offer_comparison(df):
                styles = pd.DataFrame("", index=df.index, columns=df.columns)
                for col in df.columns:
                    if col.endswith(" Mevcut"):
                        styles[col] = "background-color: #dbeafe; color: #1e3a8a; font-weight: 700; text-align:center;"
                    elif col.endswith(" Ortalama"):
                        styles[col] = "background-color: #dcfce7; color: #166534; font-weight: 700; text-align:center;"
                    elif col.endswith(" Fark"):
                        styles[col] = df[col].apply(
                            lambda x: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                            if pd.notna(x) and float(x) >= 0
                            else "background-color: #dcfce7; color: #166534; font-weight: 800; text-align:center;"
                        )
                    elif col == "Sapma Durumu":
                        styles[col] = df["Yüksek Sapma"].apply(
                            lambda is_high: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                            if bool(is_high)
                            else "background-color: #f8fafc; color: #334155; font-weight: 700; text-align:center;"
                        )
                    elif "%" in col:
                        styles[col] = "text-align:center; font-weight:700;"
                return styles

            styled_comparison = (
                comparison_display.style
                .apply(style_offer_comparison, axis=None)
                .format(formatters)
            )
            st.dataframe(
                styled_comparison,
                use_container_width=True,
                hide_index=True,
                height=420
            )

            high_deviation_rows = comparison_table[comparison_table["Yüksek Sapma"]]
            if high_deviation_rows.empty:
                st.success("Genel Toplam % açısından yüksek sapma tespit edilmedi.")
            else:
                st.warning(f"{len(high_deviation_rows)} ürün grubunda yüksek sapma tespit edildi.")

    render_consistency_section_banner(
        "Gecmis Donem Kiyaslamasi",
        "Ayni urun ve sayfa kombinasyonunda son kayda gore sert dususleri ve anomali olabilecek degisimleri tarayin.",
        tone="rose",
        eyebrow="Urun Seviyesi",
    )
    st.caption(f"Bu kıyas, hafızadaki {selected_analysis_variant.lower()} varyantı üzerinden yapılır.")
    drop_threshold = st.slider("Sert düşüş eşiği (%)", min_value=5, max_value=80, value=20, step=5, key="drop_threshold")

    reference_items = context_items.copy()
    if not reference_items.empty:
        reference_items = reference_items[reference_items[MEMORY_VARIANT_KEY_COL] != analysis_offer_key].copy()

    if not reference_items.empty and not current_item_snapshot.empty:
        key_cols = ["Ürün Grubu", "Sayfa", "Ürün Açıklaması"]
        reference_items = reference_items.sort_values("Teklif Tarihi")
        latest_reference = reference_items.groupby(key_cols, as_index=False).tail(1)
        latest_reference = latest_reference[key_cols + ["Genel Toplam", MEMORY_VARIANT_KEY_COL, "Teklif Tarihi"]].rename(
            columns={
                "Genel Toplam": "Önceki Genel Toplam",
                MEMORY_VARIANT_KEY_COL: "Önceki Analiz Kodu",
                "Teklif Tarihi": "Önceki Teklif Tarihi",
            }
        )

        comparison_df = current_item_snapshot[key_cols + ["Genel Toplam"]].rename(columns={"Genel Toplam": "Mevcut Genel Toplam"})
        comparison_df = comparison_df.merge(latest_reference, on=key_cols, how="left")
        comparison_df = comparison_df[comparison_df["Önceki Genel Toplam"].notna()].copy()

        if comparison_df.empty:
            st.info("Aynı ürün/sayfa için kıyaslanacak geçmiş kayıt bulunamadı.")
        else:
            comparison_df["Değişim %"] = (
                (comparison_df["Mevcut Genel Toplam"] - comparison_df["Önceki Genel Toplam"])
                / comparison_df["Önceki Genel Toplam"] * 100
            ).round(2)
            comparison_df["Sert Düşüş"] = comparison_df["Değişim %"] <= (-drop_threshold)
            sharp_drops = comparison_df[comparison_df["Sert Düşüş"]].sort_values("Değişim %")

            if sharp_drops.empty:
                st.success("Sert düşüş eşiğini aşan anomali tespit edilmedi.")
            else:
                st.warning(f"{len(sharp_drops)} kalemde geçmiş döneme göre sert düşüş tespit edildi.")
                st.dataframe(sharp_drops, use_container_width=True, hide_index=True, height=360)
    else:
        st.info("Geçmiş dönem kıyaslaması için bu cihaz modunda hafıza verisi henüz yok.")

    render_consistency_section_banner(
        "Teklif Hafizasina Kaydet",
        "Bu modulu ileride daha akilli hale getirmek icin mevcut teklifi secilen varyant yapisiyla hafizaya yazin.",
        tone="slate",
        eyebrow="Kayit Merkezi",
    )
    st.caption("Kaydedilen kayıtlar, ileride otomatik kıyaslamalarda cihaz dahil/hariç ayrı referans olarak ve birim fiyat geçmiş havuzunda kullanılacaktır.")

    save_variant_answer_options = ["Hayır"]
    if current_offer_has_cihazlar:
        save_variant_answer_options = ["Her İkisi", "Evet", "Hayır"]

    save_variant_answer = st.radio(
        "Bu teklifte CİHAZLAR analize dahil mi?",
        options=save_variant_answer_options,
        horizontal=True,
        key="save_offer_memory_cihaz_mode",
    )

    if st.button("Bu Teklifi Hafızaya Kaydet", key="save_offer_memory"):
        if not teklif_kodu:
            st.error("Lütfen geçerli bir Teklif Kodu girin.")
        else:
            if save_variant_answer == "Evet":
                variants_to_save = [VARIANT_DEVICE_INCLUDED]
            elif save_variant_answer == "Her İkisi":
                variants_to_save = [VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED]
            else:
                variants_to_save = [VARIANT_DEVICE_EXCLUDED]

            group_frames_to_save = [
                current_group_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_group_variants
            ]
            item_frames_to_save = [
                current_item_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_item_variants
            ]
            unit_price_frames_to_save = [
                current_unit_price_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_unit_price_variants
            ]
            group_memory_payload = pd.concat(group_frames_to_save, ignore_index=True) if group_frames_to_save else pd.DataFrame()
            item_memory_payload = pd.concat(item_frames_to_save, ignore_index=True) if item_frames_to_save else pd.DataFrame()
            unit_price_memory_payload = pd.concat(unit_price_frames_to_save, ignore_index=True) if unit_price_frames_to_save else pd.DataFrame()

            if group_memory_payload.empty and item_memory_payload.empty and unit_price_memory_payload.empty:
                st.error("Seçilen CİHAZLAR modu için kaydedilecek analiz verisi bulunamadı.")
            else:
                try:
                    if not group_memory_payload.empty:
                        upsert_memory_dataframe(
                            group_memory_path,
                            group_memory_payload,
                            key_columns=[MEMORY_VARIANT_KEY_COL, "Ürün Grubu"],
                        )
                    if not item_memory_payload.empty:
                        upsert_memory_dataframe(
                            item_memory_path,
                            item_memory_payload,
                            key_columns=[MEMORY_VARIANT_KEY_COL, "Ürün Grubu", "Sayfa", "Ürün Açıklaması"],
                        )
                    st.success("Teklif hafızaya kaydedildi: " + ", ".join(variants_to_save))
                except Exception as save_error:
                    st.error(f"Hafıza kaydı sırasında hata: {str(save_error)}")

def attach_variant_metadata(dataframe, variant_name):
    output = normalize_consistency_dataframe(dataframe)
    output[MEMORY_VARIANT_COL] = variant_name
    output[MEMORY_DEVICE_INCLUDED_COL] = variant_name == VARIANT_DEVICE_INCLUDED

    if CONS_COL_OFFER_CODE in output.columns:
        offer_codes = output[CONS_COL_OFFER_CODE].fillna("").astype(str).str.strip().tolist()
        output[MEMORY_VARIANT_KEY_COL] = [
            build_variant_offer_key(offer_code, variant_name) for offer_code in offer_codes
        ]
    else:
        output[MEMORY_VARIANT_KEY_COL] = variant_name
    return output

def recompute_group_ratio_columns(group_df):
    output = normalize_consistency_dataframe(group_df)
    numeric_columns = [
        CONS_COL_MATERIAL,
        CONS_COL_LABOR,
        CONS_COL_GGK,
        CONS_COL_TOTAL,
        CONS_COL_COUNT,
    ]
    for column in numeric_columns:
        if column in output.columns:
            output[column] = pd.to_numeric(output[column], errors="coerce").fillna(0)

    percent_column_map = [
        (CONS_COL_MATERIAL, CONS_COL_MATERIAL_PCT),
        (CONS_COL_LABOR, CONS_COL_LABOR_PCT),
        (CONS_COL_GGK, CONS_COL_GGK_PCT),
        (CONS_COL_TOTAL, CONS_COL_TOTAL_PCT),
    ]
    for amount_column, percent_column in percent_column_map:
        if amount_column not in output.columns:
            continue
        total_value = float(pd.to_numeric(output[amount_column], errors="coerce").fillna(0).sum())
        if total_value > 0:
            output[percent_column] = (
                pd.to_numeric(output[amount_column], errors="coerce").fillna(0) / total_value * 100
            ).round(2)
        else:
            output[percent_column] = 0.0
    return output

def detect_offer_has_cihazlar(group_df):
    normalized = normalize_consistency_dataframe(group_df)
    if normalized.empty or CONS_COL_GROUP not in normalized.columns or CONS_COL_MATERIAL not in normalized.columns:
        return False

    cihaz_mask = normalized[CONS_COL_GROUP].apply(is_cihazlar_group)
    if not cihaz_mask.any():
        return False

    cihaz_malzeme = pd.to_numeric(normalized.loc[cihaz_mask, CONS_COL_MATERIAL], errors="coerce").fillna(0)
    return float(cihaz_malzeme.sum()) > 0

def build_group_snapshot_variants(group_df):
    offer_df = normalize_consistency_dataframe(group_df)
    has_cihazlar = detect_offer_has_cihazlar(offer_df)
    variants = {}

    if has_cihazlar:
        variants[VARIANT_DEVICE_INCLUDED] = attach_variant_metadata(
            recompute_group_ratio_columns(offer_df.copy()),
            VARIANT_DEVICE_INCLUDED,
        )
        excluded_df = offer_df[~offer_df[CONS_COL_GROUP].apply(is_cihazlar_group)].copy()
    else:
        excluded_df = offer_df.copy()

    variants[VARIANT_DEVICE_EXCLUDED] = attach_variant_metadata(
        recompute_group_ratio_columns(excluded_df),
        VARIANT_DEVICE_EXCLUDED,
    )
    return variants, has_cihazlar

def build_item_snapshot_variants(item_df, has_cihazlar):
    offer_df = normalize_consistency_dataframe(item_df)
    variants = {}

    if has_cihazlar:
        variants[VARIANT_DEVICE_INCLUDED] = attach_variant_metadata(
            offer_df.copy(),
            VARIANT_DEVICE_INCLUDED,
        )
        excluded_df = offer_df[~offer_df[CONS_COL_GROUP].apply(is_cihazlar_group)].copy()
    else:
        excluded_df = offer_df.copy()

    variants[VARIANT_DEVICE_EXCLUDED] = attach_variant_metadata(
        excluded_df,
        VARIANT_DEVICE_EXCLUDED,
    )
    return variants

def build_unit_price_snapshot_dataframe(source_df, group_columns, scope_name):
    normalized = normalize_consistency_dataframe(source_df)
    if normalized.empty:
        return pd.DataFrame()

    required_columns = [CONS_COL_GROUP, CONS_COL_QTY, CONS_COL_MATERIAL, CONS_COL_LABOR, CONS_COL_GGK, CONS_COL_TOTAL]
    if not set(required_columns).issubset(normalized.columns):
        return pd.DataFrame()

    aggregated = normalized.groupby(group_columns, as_index=False).agg(
        **{
            CONS_COL_QTY: (CONS_COL_QTY, "sum"),
            CONS_COL_MATERIAL: (CONS_COL_MATERIAL, "sum"),
            CONS_COL_LABOR: (CONS_COL_LABOR, "sum"),
            CONS_COL_GGK: (CONS_COL_GGK, "sum"),
            CONS_COL_TOTAL: (CONS_COL_TOTAL, "sum"),
            CONS_COL_COUNT: (CONS_COL_QTY, "size"),
        }
    )
    aggregated[CONS_COL_UNIT_SCOPE] = scope_name
    if CONS_COL_UNIT not in aggregated.columns:
        aggregated[CONS_COL_UNIT] = ""

    quantity_series = pd.to_numeric(aggregated[CONS_COL_QTY], errors="coerce").replace(0, np.nan)
    aggregated[CONS_COL_MATERIAL_UNIT_PRICE] = (aggregated[CONS_COL_MATERIAL] / quantity_series).round(4)
    aggregated[CONS_COL_LABOR_UNIT_PRICE] = (aggregated[CONS_COL_LABOR] / quantity_series).round(4)
    aggregated[CONS_COL_GGK_UNIT_PRICE] = (aggregated[CONS_COL_GGK] / quantity_series).round(4)
    aggregated[CONS_COL_TOTAL_UNIT_PRICE] = (aggregated[CONS_COL_TOTAL] / quantity_series).round(4)
    aggregated = aggregated.replace([np.inf, -np.inf], np.nan)

    return aggregated.sort_values(CONS_COL_TOTAL_UNIT_PRICE, ascending=False, na_position="last").reset_index(drop=True)

def build_unit_price_snapshot_variants(dataframe, offer_code=None, offer_date=None, saved_at=None, has_cihazlar=None):
    normalized = normalize_consistency_dataframe(dataframe)
    if normalized.empty:
        return {}

    required_columns = [CONS_COL_GROUP, CONS_COL_QTY, CONS_COL_MATERIAL, CONS_COL_LABOR, CONS_COL_GGK, CONS_COL_TOTAL]
    if not set(required_columns).issubset(normalized.columns):
        return {}

    base_source = normalized[required_columns + ([CONS_COL_UNIT] if CONS_COL_UNIT in normalized.columns else [])].copy()
    if CONS_COL_UNIT not in base_source.columns:
        base_source[CONS_COL_UNIT] = ""

    base_source[CONS_COL_QTY] = pd.to_numeric(base_source[CONS_COL_QTY], errors="coerce").fillna(0.0)
    base_source[CONS_COL_UNIT] = base_source[CONS_COL_UNIT].fillna("").astype(str).str.strip()
    base_source = base_source[base_source[CONS_COL_QTY] > 0].copy()
    if base_source.empty:
        return {}

    if has_cihazlar is None:
        has_cihazlar = detect_offer_has_cihazlar(base_source[[CONS_COL_GROUP, CONS_COL_MATERIAL]].copy())

    variant_sources = {VARIANT_DEVICE_EXCLUDED: base_source.copy()}
    if has_cihazlar:
        variant_sources[VARIANT_DEVICE_INCLUDED] = base_source.copy()
        variant_sources[VARIANT_DEVICE_EXCLUDED] = base_source[
            ~base_source[CONS_COL_GROUP].apply(is_cihazlar_group)
        ].copy()

    snapshot_variants = {}
    for variant_name, variant_source in variant_sources.items():
        frames = []
        unit_specific_source = variant_source[
            (variant_source[CONS_COL_UNIT] != "") &
            (variant_source[CONS_COL_UNIT].str.lower() != "nan")
        ].copy()

        if not unit_specific_source.empty:
            frames.append(
                build_unit_price_snapshot_dataframe(
                    unit_specific_source,
                    [CONS_COL_GROUP, CONS_COL_UNIT],
                    UNIT_PRICE_SCOPE_WITH_UNIT,
                )
            )

        if not variant_source.empty:
            frames.append(
                build_unit_price_snapshot_dataframe(
                    variant_source,
                    [CONS_COL_GROUP],
                    UNIT_PRICE_SCOPE_GROUP_ONLY,
                )
            )

        snapshot_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if snapshot_df.empty:
            continue

        snapshot_df[CONS_COL_OFFER_CODE] = offer_code if offer_code is not None else ""
        snapshot_df[CONS_COL_OFFER_DATE] = pd.to_datetime(offer_date) if offer_date is not None else pd.NaT
        snapshot_df[CONS_COL_SAVED_AT] = pd.to_datetime(saved_at) if saved_at is not None else pd.NaT
        snapshot_variants[variant_name] = attach_variant_metadata(snapshot_df, variant_name)

    return snapshot_variants

def normalize_unit_price_memory(dataframe):
    output = normalize_variant_memory(dataframe)
    if output.empty and len(output.columns) == 0:
        return output

    if CONS_COL_UNIT_SCOPE not in output.columns:
        if CONS_COL_UNIT in output.columns:
            unit_text = output[CONS_COL_UNIT].fillna("").astype(str).str.strip()
            output[CONS_COL_UNIT_SCOPE] = np.where(
                (unit_text != "") & (unit_text.str.lower() != "nan"),
                UNIT_PRICE_SCOPE_WITH_UNIT,
                UNIT_PRICE_SCOPE_GROUP_ONLY,
            )
        else:
            output[CONS_COL_UNIT_SCOPE] = UNIT_PRICE_SCOPE_GROUP_ONLY

    if CONS_COL_UNIT not in output.columns:
        output[CONS_COL_UNIT] = ""

    return output

def normalize_variant_memory(dataframe):
    output = normalize_consistency_dataframe(dataframe)
    if output.empty and len(output.columns) == 0:
        return output

    if MEMORY_VARIANT_COL not in output.columns:
        output[MEMORY_VARIANT_COL] = VARIANT_DEVICE_EXCLUDED

    variant_series = output[MEMORY_VARIANT_COL].fillna("").astype(str).str.strip()
    variant_series = variant_series.replace("", VARIANT_DEVICE_EXCLUDED)
    output[MEMORY_VARIANT_COL] = variant_series.apply(
        lambda value: VARIANT_DEVICE_INCLUDED
        if "dahil" in turkce_ascii(value).lower()
        else VARIANT_DEVICE_EXCLUDED
    )

    if MEMORY_DEVICE_INCLUDED_COL in output.columns:
        output[MEMORY_DEVICE_INCLUDED_COL] = coerce_memory_bool(output[MEMORY_DEVICE_INCLUDED_COL])
    else:
        output[MEMORY_DEVICE_INCLUDED_COL] = output[MEMORY_VARIANT_COL] == VARIANT_DEVICE_INCLUDED

    if MEMORY_VARIANT_KEY_COL not in output.columns:
        if CONS_COL_OFFER_CODE in output.columns:
            offer_codes = output[CONS_COL_OFFER_CODE].fillna("").astype(str).str.strip().tolist()
            output[MEMORY_VARIANT_KEY_COL] = [
                build_variant_offer_key(offer_code, variant_name)
                for offer_code, variant_name in zip(offer_codes, output[MEMORY_VARIANT_COL].tolist())
            ]
        else:
            output[MEMORY_VARIANT_KEY_COL] = output[MEMORY_VARIANT_COL]
    return output

def expand_group_memory_variants(memory_df):
    normalized = normalize_variant_memory(memory_df)
    if normalized.empty:
        return normalized.copy(), {}

    if MEMORY_VARIANT_COL in normalized.columns and MEMORY_VARIANT_KEY_COL in normalized.columns:
        if CONS_COL_OFFER_CODE in normalized.columns:
            device_status_by_offer = (
                normalized.groupby(CONS_COL_OFFER_CODE)[MEMORY_DEVICE_INCLUDED_COL].max().astype(bool).to_dict()
            )
        else:
            device_status_by_offer = {}

        legacy_offer_keys = normalized[MEMORY_VARIANT_KEY_COL].astype(str).str.contains(r"\|", regex=True, na=False)
        if legacy_offer_keys.any():
            return normalized, {str(key).strip(): bool(value) for key, value in device_status_by_offer.items()}

    if CONS_COL_OFFER_CODE not in normalized.columns:
        return normalized, {}

    expanded_frames = []
    device_status_by_offer = {}
    for offer_code, offer_df in normalized.groupby(CONS_COL_OFFER_CODE, sort=False):
        variants, has_cihazlar = build_group_snapshot_variants(offer_df.copy())
        device_status_by_offer[str(offer_code).strip()] = has_cihazlar
        expanded_frames.extend(list(variants.values()))

    if not expanded_frames:
        return normalized, device_status_by_offer
    return pd.concat(expanded_frames, ignore_index=True), device_status_by_offer

def expand_item_memory_variants(memory_df, device_status_by_offer):
    normalized = normalize_variant_memory(memory_df)
    if normalized.empty:
        return normalized.copy()

    legacy_offer_keys = False
    if MEMORY_VARIANT_KEY_COL in normalized.columns:
        legacy_offer_keys = normalized[MEMORY_VARIANT_KEY_COL].astype(str).str.contains(r"\|", regex=True, na=False).any()
    if legacy_offer_keys:
        return normalized

    if CONS_COL_OFFER_CODE not in normalized.columns:
        return normalized

    expanded_frames = []
    for offer_code, offer_df in normalized.groupby(CONS_COL_OFFER_CODE, sort=False):
        has_cihazlar = bool(device_status_by_offer.get(str(offer_code).strip(), False))
        variants = build_item_snapshot_variants(offer_df.copy(), has_cihazlar)
        expanded_frames.extend(list(variants.values()))

    if not expanded_frames:
        return normalized
    return pd.concat(expanded_frames, ignore_index=True)

CONSISTENCY_TONE_MAP = {
    "teal": {
        "background": "linear-gradient(135deg, #ecfeff, #ccfbf1)",
        "border": "#14b8a6",
        "eyebrow": "#0f766e",
        "title": "#134e4a",
        "subtitle": "#115e59",
    },
    "blue": {
        "background": "linear-gradient(135deg, #eff6ff, #dbeafe)",
        "border": "#3b82f6",
        "eyebrow": "#1d4ed8",
        "title": "#1e3a8a",
        "subtitle": "#1d4ed8",
    },
    "amber": {
        "background": "linear-gradient(135deg, #fff7ed, #fde68a)",
        "border": "#f59e0b",
        "eyebrow": "#b45309",
        "title": "#9a3412",
        "subtitle": "#92400e",
    },
    "rose": {
        "background": "linear-gradient(135deg, #fff1f2, #ffe4e6)",
        "border": "#f43f5e",
        "eyebrow": "#be123c",
        "title": "#9f1239",
        "subtitle": "#be123c",
    },
    "slate": {
        "background": "linear-gradient(135deg, #f8fafc, #e2e8f0)",
        "border": "#64748b",
        "eyebrow": "#475569",
        "title": "#0f172a",
        "subtitle": "#334155",
    },
}

def render_consistency_page_styles():
    st.markdown(
        """
        <style>
        .consistency-hero {
            position: relative;
            margin: 6px 0 18px 0;
            padding: 30px 32px;
            border-radius: 28px;
            overflow: hidden;
            background:
                radial-gradient(circle at top right, rgba(251,191,36,0.22), transparent 32%),
                radial-gradient(circle at bottom left, rgba(45,212,191,0.18), transparent 34%),
                linear-gradient(135deg, #f8fffe 0%, #ecfeff 44%, #fff7ed 100%);
            border: 1px solid rgba(20,184,166,0.20);
            box-shadow: 0 24px 60px rgba(15,23,42,0.08);
        }
        .consistency-hero::before {
            content: "";
            position: absolute;
            inset: auto -60px -70px auto;
            width: 220px;
            height: 220px;
            background: rgba(249,115,22,0.10);
            border-radius: 50%;
            filter: blur(12px);
        }
        .consistency-eyebrow {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 8px 14px;
            border-radius: 999px;
            background: rgba(255,255,255,0.72);
            border: 1px solid rgba(148,163,184,0.20);
            color: #0f766e;
            font-size: 12px;
            font-weight: 800;
            letter-spacing: 1.6px;
            text-transform: uppercase;
            backdrop-filter: blur(8px);
        }
        .consistency-hero-title {
            margin: 16px 0 10px 0;
            font-size: 32px;
            font-weight: 900;
            letter-spacing: -0.04em;
            color: #0f172a;
            line-height: 1.05;
        }
        .consistency-hero-subtitle {
            max-width: 760px;
            font-size: 14px;
            font-weight: 600;
            color: #334155;
            line-height: 1.7;
        }
        .consistency-pill-row {
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            margin-top: 18px;
        }
        .consistency-pill {
            padding: 9px 14px;
            border-radius: 999px;
            background: rgba(255,255,255,0.82);
            border: 1px solid rgba(148,163,184,0.24);
            color: #0f172a;
            font-size: 12px;
            font-weight: 800;
            box-shadow: 0 8px 16px rgba(15,23,42,0.04);
        }
        .consistency-metric-card {
            position: relative;
            min-height: 132px;
            padding: 18px 18px 16px 18px;
            border-radius: 22px;
            border: 1px solid #dbeafe;
            box-shadow: 0 18px 36px rgba(15,23,42,0.06);
            overflow: hidden;
        }
        .consistency-metric-card::after {
            content: "";
            position: absolute;
            top: -22px;
            right: -18px;
            width: 86px;
            height: 86px;
            border-radius: 50%;
            background: rgba(255,255,255,0.38);
        }
        .consistency-metric-label {
            position: relative;
            font-size: 12px;
            font-weight: 800;
            letter-spacing: 1.2px;
            text-transform: uppercase;
            margin-bottom: 10px;
        }
        .consistency-metric-value {
            position: relative;
            font-size: 24px;
            font-weight: 900;
            line-height: 1.1;
            margin-bottom: 8px;
        }
        .consistency-metric-detail {
            position: relative;
            font-size: 13px;
            font-weight: 600;
            line-height: 1.45;
        }
        .consistency-section-banner {
            margin: 30px 0 14px 0;
            padding: 20px 22px;
            border-radius: 24px;
            border: 1px solid #dbeafe;
            box-shadow: 0 18px 34px rgba(15,23,42,0.06);
        }
        .consistency-section-eyebrow {
            font-size: 11px;
            font-weight: 900;
            letter-spacing: 1.8px;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .consistency-section-title {
            font-size: 22px;
            font-weight: 900;
            line-height: 1.1;
            margin-bottom: 6px;
        }
        .consistency-section-subtitle {
            font-size: 13px;
            font-weight: 600;
            line-height: 1.55;
        }
        .consistency-status-strip {
            margin: 10px 0 18px 0;
            padding: 14px 18px;
            border-radius: 18px;
            border: 1px solid rgba(148,163,184,0.20);
            background: linear-gradient(135deg, #ffffff, #f8fafc);
            box-shadow: 0 12px 26px rgba(15,23,42,0.05);
        }
        .consistency-status-title {
            font-size: 12px;
            font-weight: 900;
            letter-spacing: 1.5px;
            text-transform: uppercase;
            color: #0f766e;
            margin-bottom: 8px;
        }
        .consistency-status-body {
            font-size: 14px;
            font-weight: 700;
            color: #1e293b;
            line-height: 1.55;
        }
        div[data-testid="stRadio"] > label p,
        div[data-testid="stSlider"] > label p,
        div[data-testid="stMultiSelect"] > label p {
            font-weight: 800 !important;
            color: #0f172a !important;
            letter-spacing: 0.01em !important;
        }
        div[data-testid="stRadio"] [role="radiogroup"] {
            gap: 10px;
            padding: 10px;
            border-radius: 20px;
            border: 1px solid #dbe4ee;
            background: linear-gradient(135deg, #f8fafc, #ffffff);
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.9);
        }
        div[data-testid="stRadio"] [role="radiogroup"] > label {
            border: 1px solid #cbd5e1;
            border-radius: 999px;
            padding: 6px 14px;
            background: #ffffff;
            box-shadow: 0 6px 14px rgba(15,23,42,0.04);
        }
        div[data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 10px;
            padding: 10px;
            border-radius: 20px;
            background: linear-gradient(135deg, #f8fafc, #ffffff);
            border: 1px solid #dbeafe;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.82);
        }
        div[data-testid="stTabs"] [data-baseweb="tab"] {
            height: 46px;
            border-radius: 14px;
            border: 1px solid #dbe4ee;
            background: #ffffff;
            color: #334155;
            font-weight: 800;
            padding: 0 18px;
        }
        div[data-testid="stTabs"] [data-baseweb="tab"][aria-selected="true"] {
            background: linear-gradient(135deg, #0f766e, #0ea5e9);
            color: #ffffff;
            border-color: transparent;
            box-shadow: 0 12px 24px rgba(14,165,233,0.24);
        }
        div[data-testid="stTabs"] [data-baseweb="tab-highlight"] {
            background: transparent;
        }
        div[data-testid="stAlert"] {
            border-radius: 18px;
            border: 1px solid rgba(148,163,184,0.24);
            box-shadow: 0 12px 24px rgba(15,23,42,0.05);
        }
        div[data-testid="stButton"] button {
            min-height: 50px;
            border-radius: 16px;
            border: 1px solid #0f766e;
            background: linear-gradient(135deg, #0f766e, #14b8a6);
            color: #ffffff;
            font-weight: 900;
            letter-spacing: 0.02em;
            box-shadow: 0 16px 30px rgba(20,184,166,0.22);
        }
        div[data-testid="stButton"] button:hover {
            border-color: #115e59;
            box-shadow: 0 18px 34px rgba(20,184,166,0.28);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

def render_consistency_hero(source_file_name, teklif_kodu, selected_sheets, current_offer_has_cihazlar):
    safe_offer_code = escape(str(teklif_kodu or "Bilinmiyor"))
    safe_file_name = escape(str(source_file_name or "Dosya"))
    sheet_count = len(selected_sheets or [])
    sheet_label = f"{sheet_count} sayfa secili"
    cihaz_label = "Cihazli teklif akisi" if current_offer_has_cihazlar else "Cihaz haric teklif akisi"

    st.markdown(
        f"""
        <div class="consistency-hero">
            <div class="consistency-eyebrow">Karsilastirma Merkezi</div>
            <div class="consistency-hero-title">Tutarlilik, gecmis hafiza ve birim fiyat kiyasi tek ekranda.</div>
            <div class="consistency-hero-subtitle">
                Bu alan mevcut teklifi gecmis dosyalarla ayni baglamda okur, Cihazlar dahil ve haric varyantlarini ayri ele alir
                ve kritik sapmalari daha hizli fark etmeniz icin akisi tek merkezde toplar.
            </div>
            <div class="consistency-pill-row">
                <div class="consistency-pill">Teklif Kodu: {safe_offer_code}</div>
                <div class="consistency-pill">Kaynak Dosya: {safe_file_name}</div>
                <div class="consistency-pill">{escape(sheet_label)}</div>
                <div class="consistency-pill">{escape(cihaz_label)}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_consistency_metric_cards(cards):
    if not cards:
        return

    columns = st.columns(len(cards))
    for column, card in zip(columns, cards):
        tone_name = card.get("tone", "slate")
        tone = CONSISTENCY_TONE_MAP.get(tone_name, CONSISTENCY_TONE_MAP["slate"])
        with column:
            st.markdown(
                f"""
                <div class="consistency-metric-card" style="background:{tone['background']}; border-color:{tone['border']};">
                    <div class="consistency-metric-label" style="color:{tone['eyebrow']};">{escape(str(card.get('label', '')))}</div>
                    <div class="consistency-metric-value" style="color:{tone['title']};">{escape(str(card.get('value', '')))}</div>
                    <div class="consistency-metric-detail" style="color:{tone['subtitle']};">{escape(str(card.get('detail', '')))}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

def render_consistency_section_banner(title, subtitle="", tone="teal", eyebrow="Analiz Bolumu"):
    palette = CONSISTENCY_TONE_MAP.get(tone, CONSISTENCY_TONE_MAP["teal"])
    st.markdown(
        f"""
        <div class="consistency-section-banner" style="background:{palette['background']}; border-color:{palette['border']};">
            <div class="consistency-section-eyebrow" style="color:{palette['eyebrow']};">{escape(str(eyebrow))}</div>
            <div class="consistency-section-title" style="color:{palette['title']};">{escape(str(title))}</div>
            <div class="consistency-section-subtitle" style="color:{palette['subtitle']};">{escape(str(subtitle))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_consistency_status_strip(title, body):
    st.markdown(
        f"""
        <div class="consistency-status-strip">
            <div class="consistency-status-title">{escape(str(title))}</div>
            <div class="consistency-status-body">{escape(str(body))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_consistency_module(all_df, summary_data, selected_sheets, source_file_name):
    render_consistency_page_styles()

    normalized_all_df = normalize_consistency_dataframe(all_df)
    normalized_summary = normalize_consistency_dataframe(summary_data)

    teklif_kodu = derive_offer_code_from_outputs(
        all_df,
        selected_sheets=selected_sheets,
        fallback_name=source_file_name,
    )
    teklif_tarihi = pd.Timestamp.now().normalize()
    kayit_zamani = pd.Timestamp.now()

    render_consistency_hero(
        source_file_name=source_file_name,
        teklif_kodu=teklif_kodu,
        selected_sheets=selected_sheets,
        current_offer_has_cihazlar=detect_offer_has_cihazlar(normalized_summary),
    )

    group_memory_path, item_memory_path, _ = get_offer_memory_paths()
    unit_price_memory_path = get_unit_price_memory_path()
    history_groups_raw = normalize_consistency_dataframe(load_memory_dataframe(group_memory_path))
    history_items_raw = normalize_consistency_dataframe(load_memory_dataframe(item_memory_path))
    history_unit_prices_raw = normalize_unit_price_memory(load_memory_dataframe(unit_price_memory_path))
    history_groups, device_status_by_offer = expand_group_memory_variants(history_groups_raw)
    history_items = expand_item_memory_variants(history_items_raw, device_status_by_offer)
    history_unit_prices = normalize_unit_price_memory(history_unit_prices_raw)

    current_group_base = normalized_summary[
        [
            CONS_COL_GROUP,
            CONS_COL_MATERIAL,
            CONS_COL_LABOR,
            CONS_COL_GGK,
            CONS_COL_TOTAL,
            CONS_COL_COUNT,
            CONS_COL_MATERIAL_PCT,
            CONS_COL_LABOR_PCT,
            CONS_COL_GGK_PCT,
            CONS_COL_TOTAL_PCT,
        ]
    ].copy()
    current_group_base[CONS_COL_OFFER_CODE] = teklif_kodu
    current_group_base[CONS_COL_OFFER_DATE] = pd.to_datetime(teklif_tarihi)
    current_group_base[CONS_COL_SELECTED_SHEETS] = " | ".join(selected_sheets)
    current_group_base[CONS_COL_SAVED_AT] = kayit_zamani

    current_item_base = normalized_all_df[
        [
            CONS_COL_GROUP,
            CONS_COL_PAGE,
            CONS_COL_DESC,
            CONS_COL_MATERIAL,
            CONS_COL_LABOR,
            CONS_COL_GGK,
            CONS_COL_TOTAL,
        ]
    ].copy()
    current_item_base = current_item_base.groupby(
        [CONS_COL_GROUP, CONS_COL_PAGE, CONS_COL_DESC], as_index=False
    ).agg(
        {
            CONS_COL_MATERIAL: "sum",
            CONS_COL_LABOR: "sum",
            CONS_COL_GGK: "sum",
            CONS_COL_TOTAL: "sum",
        }
    )
    current_item_base[CONS_COL_OFFER_CODE] = teklif_kodu
    current_item_base[CONS_COL_OFFER_DATE] = pd.to_datetime(teklif_tarihi)
    current_item_base[CONS_COL_SAVED_AT] = kayit_zamani

    current_group_variants, current_offer_has_cihazlar = build_group_snapshot_variants(current_group_base)
    current_item_variants = build_item_snapshot_variants(current_item_base, current_offer_has_cihazlar)
    current_unit_price_variants = build_unit_price_snapshot_variants(
        normalized_all_df,
        offer_code=teklif_kodu,
        offer_date=teklif_tarihi,
        saved_at=kayit_zamani,
        has_cihazlar=current_offer_has_cihazlar,
    )

    cihaz_mask = current_group_base[CONS_COL_GROUP].apply(is_cihazlar_group)
    cihaz_malzeme_toplami = sum_numeric_column(current_group_base.loc[cihaz_mask].copy(), CONS_COL_MATERIAL)

    analysis_variant_options = [VARIANT_DEVICE_EXCLUDED]
    if current_offer_has_cihazlar:
        analysis_variant_options = [VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED]

    selected_analysis_variant = st.radio(
        "CİHAZLAR analiz modu",
        options=analysis_variant_options,
        horizontal=True,
        key="cihazlar_analysis_variant",
    )

    current_group_snapshot = current_group_variants[selected_analysis_variant].copy()
    current_item_snapshot = current_item_variants[selected_analysis_variant].copy()
    analysis_offer_key = build_variant_offer_key(teklif_kodu, selected_analysis_variant)

    context_groups = history_groups.copy()
    context_items = history_items.copy()
    context_unit_prices = history_unit_prices.copy()
    if not context_groups.empty:
        context_groups = context_groups[context_groups[MEMORY_VARIANT_COL] == selected_analysis_variant].copy()
    if not context_items.empty:
        context_items = context_items[context_items[MEMORY_VARIANT_COL] == selected_analysis_variant].copy()
    if not context_unit_prices.empty:
        context_unit_prices = context_unit_prices[context_unit_prices[MEMORY_VARIANT_COL] == selected_analysis_variant].copy()

    history_offer_count = context_groups[MEMORY_VARIANT_KEY_COL].nunique() if not context_groups.empty else 0
    unit_price_offer_count = context_unit_prices[MEMORY_VARIANT_KEY_COL].nunique() if not context_unit_prices.empty else 0
    render_consistency_metric_cards(
        [
            {
                "label": "Aktif Analiz Modu",
                "value": selected_analysis_variant,
                "detail": "Bu varyant uzerinden tum kiyaslar yapiliyor.",
                "tone": "teal",
            },
            {
                "label": "Cihaz Malzeme Toplami",
                "value": format_currency_display(cihaz_malzeme_toplami),
                "detail": "Mevcut teklif icindeki cihaz malzeme agirligi.",
                "tone": "amber",
            },
            {
                "label": "Gecmis Teklif Havuzu",
                "value": format_integer_display(history_offer_count),
                "detail": f"Grup kiyasi icin {len(context_groups)} hafiza satiri kullaniliyor.",
                "tone": "blue",
            },
            {
                "label": "Birim Fiyat Hafizasi",
                "value": format_integer_display(unit_price_offer_count),
                "detail": f"Birim fiyat tarafinda {len(context_unit_prices)} satir referans bulundu.",
                "tone": "slate",
            },
        ]
    )
    if current_offer_has_cihazlar:
        render_consistency_status_strip(
            "Cihazli teklif akisi aktif",
            "Bu teklifte CİHAZLAR malzemesi bulundu. Kıyas ve hafıza kayıtları cihaz dahil ve cihaz hariç olarak ayrı yürür.",
        )
    else:
        render_consistency_status_strip(
            "Cihaz haric referans havuzu aktif",
            "Bu teklifte CİHAZLAR malzeme toplamı 0 olduğu için cihaz hariç referans havuzu kullanılacaktır.",
        )

    render_consistency_section_banner(
        "Mevcut Teklif Yuzdelik Dagilim",
        "Malzeme, iscilik ve GGK katmanlarini tek bakista okuyun. Bu alan teklifin temel denge fotografini verir.",
        tone="teal",
        eyebrow="Mevcut Teklif",
    )
    toplam_malzeme = sum_numeric_column(current_item_snapshot, CONS_COL_MATERIAL)
    toplam_iscilik = sum_numeric_column(current_item_snapshot, CONS_COL_LABOR)
    toplam_ggk = sum_numeric_column(current_item_snapshot, CONS_COL_GGK)
    toplam_genel = sum_numeric_column(current_item_snapshot, CONS_COL_TOTAL)
    mevcut_dagilim_df = pd.DataFrame(
        [
            {"Kalem": "Malzeme", "Tutar": toplam_malzeme},
            {"Kalem": "İşçilik", "Tutar": toplam_iscilik},
            {"Kalem": "GGK", "Tutar": toplam_ggk},
        ]
    )
    if toplam_genel > 0:
        mevcut_dagilim_df["Teklif İçindeki %"] = (mevcut_dagilim_df["Tutar"] / toplam_genel * 100).round(2)
    else:
        mevcut_dagilim_df["Teklif İçindeki %"] = 0.0
    create_sortable_numeric_table(
        mevcut_dagilim_df,
        height=180,
        currency_cols=["Tutar"],
        percent_cols=["Teklif İçindeki %"],
    )

    render_consistency_section_banner(
        "Teklifler Arasi Urun Grubu Yuzde Kiyasi",
        "Mevcut teklif dagilimini gecmis ortalamalarla yan yana okuyun ve sapma olan alanlari hizli sekilde tespit edin.",
        tone="blue",
        eyebrow="Gecmis Ortalama",
    )
    if current_offer_has_cihazlar:
        st.caption("Bu bölümde CİHAZLAR için cihaz dahil ve cihaz hariç ortalama kıyasları ayrı hesaplanır.")
    else:
        st.caption(f"Ürün grubu % kıyası, hafızadaki {selected_analysis_variant.lower()} varyantı üzerinden yapılır.")

    ratio_cols = [CONS_COL_TOTAL_PCT, CONS_COL_MATERIAL_PCT, CONS_COL_LABOR_PCT, CONS_COL_GGK_PCT]
    compare_mode = st.radio(
        "Kıyas Kaynağı",
        options=["Tüm geçmiş teklifler ortalaması", "Seçilen teklifler"],
        horizontal=True,
        key="offer_compare_mode",
    )
    deviation_threshold = st.slider(
        "Genel Toplam % sapma eşiği (puan)",
        min_value=1,
        max_value=50,
        value=10,
        step=1,
        key="offer_ratio_deviation_threshold",
    )

    def style_offer_comparison(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in df.columns:
            if col.endswith(" Mevcut"):
                styles[col] = "background-color: #dbeafe; color: #1e3a8a; font-weight: 700; text-align:center;"
            elif col.endswith(" Ortalama"):
                styles[col] = "background-color: #dcfce7; color: #166534; font-weight: 700; text-align:center;"
            elif col.endswith(" Fark"):
                styles[col] = df[col].apply(
                    lambda x: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                    if pd.notna(x) and float(x) >= 0
                    else "background-color: #dcfce7; color: #166534; font-weight: 800; text-align:center;"
                )
            elif col == "Sapma Durumu":
                styles[col] = df["Yüksek Sapma"].apply(
                    lambda is_high: "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                    if bool(is_high)
                    else "background-color: #f8fafc; color: #334155; font-weight: 700; text-align:center;"
                )
            elif "%" in col:
                styles[col] = "text-align:center; font-weight:700;"
        return styles

    def render_average_compare_for_variant(variant_name):
        variant_slug = turkce_ascii(str(variant_name)).lower().replace(" ", "_")
        variant_offer_key = build_variant_offer_key(teklif_kodu, variant_name)
        variant_current_group = current_group_variants.get(variant_name, pd.DataFrame()).copy()
        variant_historical_groups = history_groups.copy()
        if not variant_historical_groups.empty:
            variant_historical_groups = variant_historical_groups[
                variant_historical_groups[MEMORY_VARIANT_COL] == variant_name
            ].copy()
            variant_historical_groups = variant_historical_groups[
                variant_historical_groups[MEMORY_VARIANT_KEY_COL] != variant_offer_key
            ].copy()

        st.caption(f"{variant_name} varyantı hesaplaması")

        if variant_historical_groups.empty:
            st.info(f"{variant_name} için geçmiş teklif verisi henüz yok.")
            return
        if variant_current_group.empty:
            st.info(f"{variant_name} için mevcut teklifte kıyaslanacak ürün grubu kalmadı.")
            return

        offer_dates = variant_historical_groups.groupby(MEMORY_VARIANT_KEY_COL, as_index=False)[CONS_COL_OFFER_DATE].max()
        offer_meta = variant_historical_groups.groupby(MEMORY_VARIANT_KEY_COL, as_index=False).agg(
            {
                CONS_COL_OFFER_CODE: "first",
                MEMORY_VARIANT_COL: "first",
            }
        )
        historical_offer_table = variant_historical_groups.pivot_table(
            index=MEMORY_VARIANT_KEY_COL,
            columns=CONS_COL_GROUP,
            values=CONS_COL_TOTAL_PCT,
            aggfunc="sum",
            fill_value=0.0
        ).reset_index()
        historical_offer_table = historical_offer_table.merge(offer_meta, on=MEMORY_VARIANT_KEY_COL, how="left")
        historical_offer_table = historical_offer_table.merge(offer_dates, on=MEMORY_VARIANT_KEY_COL, how="left")

        ordered_cols = [MEMORY_VARIANT_KEY_COL, CONS_COL_OFFER_CODE, MEMORY_VARIANT_COL, CONS_COL_OFFER_DATE] + [
            col for col in historical_offer_table.columns
            if col not in [MEMORY_VARIANT_KEY_COL, CONS_COL_OFFER_CODE, MEMORY_VARIANT_COL, CONS_COL_OFFER_DATE]
        ]

        st.markdown("#### Geçmiş Tekliflerin Ürün Grubu Yüzdeleri")
        st.dataframe(
            historical_offer_table[ordered_cols].sort_values(CONS_COL_OFFER_DATE, ascending=False),
            use_container_width=True,
            hide_index=True,
            height=260
        )

        if compare_mode == "Seçilen teklifler":
            offer_options = historical_offer_table[MEMORY_VARIANT_KEY_COL].astype(str).tolist()
            offer_display_map = {}
            for _, row in historical_offer_table.iterrows():
                offer_key = str(row.get(MEMORY_VARIANT_KEY_COL, "")).strip()
                offer_code_text = str(row.get(CONS_COL_OFFER_CODE, "")).strip() or offer_key
                offer_date = pd.to_datetime(row.get(CONS_COL_OFFER_DATE), errors="coerce")
                if pd.notna(offer_date):
                    offer_display_map[offer_key] = f"{offer_code_text} ({offer_date.strftime('%d.%m.%Y')})"
                else:
                    offer_display_map[offer_key] = offer_code_text
            selected_offers = st.multiselect(
                f"Kıyas için teklif seçin ({variant_name})",
                options=offer_options,
                default=offer_options[:1] if offer_options else [],
                format_func=lambda option, mapping=offer_display_map: mapping.get(str(option), str(option)),
                key=f"selected_offer_codes_{variant_slug}",
            )
            benchmark_groups = variant_historical_groups[
                variant_historical_groups[MEMORY_VARIANT_KEY_COL].astype(str).isin(selected_offers)
            ].copy()
        else:
            benchmark_groups = variant_historical_groups.copy()

        if benchmark_groups.empty:
            st.warning(f"{variant_name} için seçilen kıyas kaynağında veri yok.")
            return

        benchmark_avg = benchmark_groups.groupby(CONS_COL_GROUP)[ratio_cols].mean().reset_index()
        benchmark_count = benchmark_groups.groupby(CONS_COL_GROUP)[MEMORY_VARIANT_KEY_COL].nunique().reset_index(name="Referans Teklif Sayısı")

        comparison_table = variant_current_group[[CONS_COL_GROUP] + ratio_cols].merge(
            benchmark_avg,
            on=CONS_COL_GROUP,
            how="left",
            suffixes=(" Mevcut", " Ortalama")
        )
        comparison_table = comparison_table.merge(benchmark_count, on=CONS_COL_GROUP, how="left")
        comparison_table["Referans Teklif Sayısı"] = comparison_table["Referans Teklif Sayısı"].fillna(0).astype(int)

        global_avg = benchmark_groups[ratio_cols].mean()
        for ratio_col in ratio_cols:
            comparison_table[f"{ratio_col} Ortalama"] = comparison_table[f"{ratio_col} Ortalama"].fillna(global_avg.get(ratio_col, np.nan))
            comparison_table[f"{ratio_col} Fark"] = (
                comparison_table[f"{ratio_col} Mevcut"] - comparison_table[f"{ratio_col} Ortalama"]
            ).round(2)

        comparison_table["Yüksek Sapma"] = comparison_table[f"{CONS_COL_TOTAL_PCT} Fark"].abs() >= deviation_threshold
        comparison_table = comparison_table.sort_values(f"{CONS_COL_TOTAL_PCT} Fark", key=lambda s: s.abs(), ascending=False)

        st.markdown(f"#### Mevcut Teklif vs Kıyas Kaynağı (Ortalama) - {variant_name}")
        comparison_display = comparison_table.copy()
        comparison_display["Sapma Durumu"] = np.where(comparison_display["Yüksek Sapma"], "Yüksek Sapma", "Normal")

        ordered_cols = [CONS_COL_GROUP]
        for ratio_col in ratio_cols:
            ordered_cols.extend([f"{ratio_col} Mevcut", f"{ratio_col} Ortalama", f"{ratio_col} Fark"])
        ordered_cols.extend(["Referans Teklif Sayısı", "Sapma Durumu", "Yüksek Sapma"])
        comparison_display = comparison_display[ordered_cols]

        percent_cols = [col for col in comparison_display.columns if "%" in col]
        formatters = {}
        for col in percent_cols:
            if col.endswith(" Fark"):
                formatters[col] = lambda v: f"{float(v):+.2f} puan" if pd.notna(v) else ""
            else:
                formatters[col] = lambda v: f"{float(v):.2f}%" if pd.notna(v) else ""

        styled_comparison = (
            comparison_display.style
            .apply(style_offer_comparison, axis=None)
            .format(formatters)
        )
        st.dataframe(styled_comparison, use_container_width=True, hide_index=True, height=420)

        high_deviation_rows = comparison_table[comparison_table["Yüksek Sapma"]]
        if high_deviation_rows.empty:
            st.success(f"{variant_name} için Genel Toplam % açısından yüksek sapma tespit edilmedi.")
        else:
            st.warning(f"{variant_name} için {len(high_deviation_rows)} ürün grubunda yüksek sapma tespit edildi.")

    if current_offer_has_cihazlar:
        variant_tabs = st.tabs([VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED])
        for tab, variant_name in zip(variant_tabs, [VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED]):
            with tab:
                render_average_compare_for_variant(variant_name)
    else:
        render_average_compare_for_variant(VARIANT_DEVICE_EXCLUDED)

    render_consistency_section_banner(
        "Birim Fiyat Analizi ve Gecmis Kiyas",
        "Hem urun grubu + birim hem de sadece urun grubu bazinda gecmis teklif ortalamalariyla daha net birim fiyat okuması alin.",
        tone="amber",
        eyebrow="Birim Fiyat",
    )
    st.caption(f"Bu bölüm, hafızadaki {selected_analysis_variant.lower()} varyantı için birim fiyatları geçmiş dosya ortalamalarıyla karşılaştırır.")

    current_unit_price_snapshot = current_unit_price_variants.get(selected_analysis_variant, pd.DataFrame()).copy()
    reference_unit_prices = context_unit_prices.copy()
    if not reference_unit_prices.empty:
        reference_unit_prices = reference_unit_prices[
            reference_unit_prices[MEMORY_VARIANT_KEY_COL] != analysis_offer_key
        ].copy()

    unit_price_compare_mode = st.radio(
        "Birim fiyat kıyas kaynağı",
        options=["Tüm geçmiş teklifler ortalaması", "Seçilen teklifler"],
        horizontal=True,
        key="unit_price_compare_mode",
    )
    unit_price_deviation_threshold = st.slider(
        "Genel Toplam birim fiyat sapma eşiği (%)",
        min_value=1,
        max_value=100,
        value=15,
        step=1,
        key="unit_price_deviation_threshold",
    )

    unit_price_metric_cols = [
        CONS_COL_MATERIAL_UNIT_PRICE,
        CONS_COL_LABOR_UNIT_PRICE,
        CONS_COL_GGK_UNIT_PRICE,
        CONS_COL_TOTAL_UNIT_PRICE,
    ]
    unit_price_currency_cols = [
        CONS_COL_MATERIAL,
        CONS_COL_LABOR,
        CONS_COL_GGK,
        CONS_COL_TOTAL,
    ] + unit_price_metric_cols

    def style_unit_price_comparison(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        reference_counts = pd.to_numeric(df.get("Referans Teklif Sayisi"), errors="coerce").fillna(0)
        for col in df.columns:
            if col.endswith(" Mevcut"):
                styles[col] = "background-color: #dbeafe; color: #1e3a8a; font-weight: 700; text-align:center;"
            elif col.endswith(" Ortalama"):
                styles[col] = "background-color: #dcfce7; color: #166534; font-weight: 700; text-align:center;"
            elif col.endswith(" Fark %"):
                styles[col] = [
                    "background-color: #f8fafc; color: #475569; font-weight: 700; text-align:center;"
                    if ref_count <= 0 or pd.isna(value)
                    else "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                    if float(value) >= 0
                    else "background-color: #dcfce7; color: #166534; font-weight: 800; text-align:center;"
                    for value, ref_count in zip(df[col].tolist(), reference_counts.tolist())
                ]
            elif col == "Sapma Durumu":
                styles[col] = [
                    "background-color: #f8fafc; color: #475569; font-weight: 700; text-align:center;"
                    if value == "Referans Yok"
                    else "background-color: #fee2e2; color: #991b1b; font-weight: 800; text-align:center;"
                    if value == "Yüksek Sapma"
                    else "background-color: #ecfccb; color: #3f6212; font-weight: 800; text-align:center;"
                    for value in df[col].tolist()
                ]
        return styles

    def render_unit_price_history_scope(scope_name, scope_title, include_unit):
        if CONS_COL_UNIT_SCOPE not in current_unit_price_snapshot.columns:
            st.info(f"{scope_title} için mevcut teklifte birim fiyat analizi yapılacak veri bulunamadı.")
            return

        scope_current = current_unit_price_snapshot[
            current_unit_price_snapshot[CONS_COL_UNIT_SCOPE] == scope_name
        ].copy()
        if CONS_COL_UNIT_SCOPE in reference_unit_prices.columns:
            scope_history = reference_unit_prices[
                reference_unit_prices[CONS_COL_UNIT_SCOPE] == scope_name
            ].copy()
        else:
            scope_history = pd.DataFrame()

        if scope_current.empty:
            st.info(f"{scope_title} için mevcut teklifte hesaplanacak birim fiyat satırı bulunamadı.")
            return

        base_columns = [CONS_COL_GROUP]
        if include_unit:
            base_columns.append(CONS_COL_UNIT)
        current_display_cols = base_columns + [
            CONS_COL_QTY,
            CONS_COL_COUNT,
            CONS_COL_MATERIAL,
            CONS_COL_MATERIAL_UNIT_PRICE,
            CONS_COL_LABOR,
            CONS_COL_LABOR_UNIT_PRICE,
            CONS_COL_GGK,
            CONS_COL_GGK_UNIT_PRICE,
            CONS_COL_TOTAL,
            CONS_COL_TOTAL_UNIT_PRICE,
        ]

        st.markdown(f"#### {scope_title} - Mevcut Teklif")
        create_sortable_numeric_table(
            scope_current[current_display_cols],
            height=320,
            currency_cols=unit_price_currency_cols,
            integer_cols=[CONS_COL_QTY, CONS_COL_COUNT],
        )

        if scope_history.empty:
            st.info(f"{scope_title} için geçmiş dosya referansı henüz yok.")
            return

        history_offer_meta = scope_history.groupby(MEMORY_VARIANT_KEY_COL, as_index=False).agg(
            {
                CONS_COL_OFFER_CODE: "first",
                CONS_COL_OFFER_DATE: "max",
            }
        ).sort_values(CONS_COL_OFFER_DATE, ascending=False)

        if unit_price_compare_mode == "Seçilen teklifler":
            offer_options = history_offer_meta[MEMORY_VARIANT_KEY_COL].astype(str).tolist()
            offer_display_map = {}
            for _, row in history_offer_meta.iterrows():
                offer_key = str(row.get(MEMORY_VARIANT_KEY_COL, "")).strip()
                offer_code_text = str(row.get(CONS_COL_OFFER_CODE, "")).strip() or offer_key
                offer_date = pd.to_datetime(row.get(CONS_COL_OFFER_DATE), errors="coerce")
                if pd.notna(offer_date):
                    offer_display_map[offer_key] = f"{offer_code_text} ({offer_date.strftime('%d.%m.%Y')})"
                else:
                    offer_display_map[offer_key] = offer_code_text
            selected_offers = st.multiselect(
                f"Kıyas için teklif seçin - {scope_title}",
                options=offer_options,
                default=offer_options[:1] if offer_options else [],
                format_func=lambda option, mapping=offer_display_map: mapping.get(str(option), str(option)),
                key=f"unit_price_offer_select_{turkce_ascii(scope_name).lower().replace(' ', '_')}",
            )
            benchmark_source = scope_history[
                scope_history[MEMORY_VARIANT_KEY_COL].astype(str).isin(selected_offers)
            ].copy()
        else:
            benchmark_source = scope_history.copy()

        if benchmark_source.empty:
            st.warning(f"{scope_title} için seçilen kıyas kaynağında veri yok.")
            return

        key_cols = [CONS_COL_GROUP, CONS_COL_UNIT] if include_unit else [CONS_COL_GROUP]
        benchmark_avg = benchmark_source.groupby(key_cols, as_index=False)[unit_price_metric_cols].mean()
        benchmark_count = benchmark_source.groupby(key_cols)[MEMORY_VARIANT_KEY_COL].nunique().reset_index(name="Referans Teklif Sayisi")

        comparison_df = scope_current[key_cols + [CONS_COL_QTY, CONS_COL_COUNT] + unit_price_metric_cols].merge(
            benchmark_avg,
            on=key_cols,
            how="left",
            suffixes=(" Mevcut", " Ortalama"),
        )
        comparison_df = comparison_df.merge(benchmark_count, on=key_cols, how="left")
        comparison_df["Referans Teklif Sayisi"] = comparison_df["Referans Teklif Sayisi"].fillna(0).astype(int)

        total_average_col = f"{CONS_COL_TOTAL_UNIT_PRICE} Ortalama"
        total_current_col = f"{CONS_COL_TOTAL_UNIT_PRICE} Mevcut"
        total_diff_col = f"{CONS_COL_TOTAL_UNIT_PRICE} Fark %"
        average_series = pd.to_numeric(comparison_df[total_average_col], errors="coerce")
        current_series = pd.to_numeric(comparison_df[total_current_col], errors="coerce")
        comparison_df[total_diff_col] = np.where(
            average_series > 0,
            ((current_series - average_series) / average_series * 100).round(2),
            np.nan,
        )
        comparison_df["Yüksek Sapma"] = comparison_df[total_diff_col].abs() >= unit_price_deviation_threshold
        comparison_df["Sapma Durumu"] = np.where(
            comparison_df["Referans Teklif Sayisi"] <= 0,
            "Referans Yok",
            np.where(comparison_df["Yüksek Sapma"], "Yüksek Sapma", "Normal"),
        )
        comparison_df = comparison_df.sort_values(total_diff_col, key=lambda s: s.abs(), ascending=False, na_position="last")

        ordered_cols = key_cols + [CONS_COL_QTY, CONS_COL_COUNT]
        for metric_col in unit_price_metric_cols:
            ordered_cols.extend([f"{metric_col} Mevcut", f"{metric_col} Ortalama"])
        ordered_cols.extend([total_diff_col, "Referans Teklif Sayisi", "Sapma Durumu", "Yüksek Sapma"])
        comparison_display = comparison_df[ordered_cols].copy()

        formatters = {
            total_diff_col: lambda value: f"{float(value):+.2f}%" if pd.notna(value) else "",
        }
        for metric_col in unit_price_metric_cols:
            for suffix in [" Mevcut", " Ortalama"]:
                formatters[f"{metric_col}{suffix}"] = format_currency_display
        formatters[CONS_COL_QTY] = format_integer_display
        formatters[CONS_COL_COUNT] = format_integer_display
        formatters["Referans Teklif Sayisi"] = format_integer_display

        st.markdown(f"#### {scope_title} - Geçmiş Dosya Ortalaması Kıyas")
        st.caption(f"Kıyas havuzu: {benchmark_source[MEMORY_VARIANT_KEY_COL].nunique()} teklif")
        st.dataframe(
            comparison_display.style.apply(style_unit_price_comparison, axis=None).format(formatters),
            use_container_width=True,
            hide_index=True,
            height=420,
        )

        matched_rows = comparison_df[comparison_df["Referans Teklif Sayisi"] > 0]
        if matched_rows.empty:
            st.info(f"{scope_title} için eşleşen geçmiş teklif kaydı bulunamadı.")
        elif matched_rows["Yüksek Sapma"].any():
            st.warning(f"{scope_title} için {int(matched_rows['Yüksek Sapma'].sum())} satırda yüksek birim fiyat sapması tespit edildi.")
        else:
            st.success(f"{scope_title} için eşleşen geçmiş kayıtlarla kıyaslandığında yüksek birim fiyat sapması bulunmadı.")

    unit_price_tabs = st.tabs(["Ürün Grubu + Birim", "Sadece Ürün Grubu"])
    with unit_price_tabs[0]:
        render_unit_price_history_scope(UNIT_PRICE_SCOPE_WITH_UNIT, "Ürün Grubu + Birim", include_unit=True)
    with unit_price_tabs[1]:
        render_unit_price_history_scope(UNIT_PRICE_SCOPE_GROUP_ONLY, "Sadece Ürün Grubu", include_unit=False)

    render_consistency_section_banner(
        "Gecmis Donem Kiyaslamasi",
        "Ayni urun ve sayfa kombinasyonunda son kayda gore sert dususleri ve anomali olabilecek degisimleri tarayin.",
        tone="rose",
        eyebrow="Urun Seviyesi",
    )
    st.caption(f"Bu kıyas, hafızadaki {selected_analysis_variant.lower()} varyantı üzerinden yapılır.")
    drop_threshold = st.slider("Sert düşüş eşiği (%)", min_value=5, max_value=80, value=20, step=5, key="drop_threshold")

    reference_items = context_items.copy()
    if not reference_items.empty:
        reference_items = reference_items[reference_items[MEMORY_VARIANT_KEY_COL] != analysis_offer_key].copy()

    if not reference_items.empty and not current_item_snapshot.empty:
        key_cols = [CONS_COL_GROUP, CONS_COL_PAGE, CONS_COL_DESC]
        reference_items = reference_items.sort_values(CONS_COL_OFFER_DATE)
        latest_reference = reference_items.groupby(key_cols, as_index=False).tail(1)
        latest_reference = latest_reference[key_cols + [CONS_COL_TOTAL, MEMORY_VARIANT_KEY_COL, CONS_COL_OFFER_DATE]].rename(
            columns={
                CONS_COL_TOTAL: "Önceki Genel Toplam",
                MEMORY_VARIANT_KEY_COL: "Önceki Analiz Kodu",
                CONS_COL_OFFER_DATE: "Önceki Teklif Tarihi",
            }
        )

        comparison_df = current_item_snapshot[key_cols + [CONS_COL_TOTAL]].rename(columns={CONS_COL_TOTAL: "Mevcut Genel Toplam"})
        comparison_df = comparison_df.merge(latest_reference, on=key_cols, how="left")
        comparison_df = comparison_df[comparison_df["Önceki Genel Toplam"].notna()].copy()

        if comparison_df.empty:
            st.info("Aynı ürün/sayfa için kıyaslanacak geçmiş kayıt bulunamadı.")
        else:
            comparison_df["Değişim %"] = (
                (comparison_df["Mevcut Genel Toplam"] - comparison_df["Önceki Genel Toplam"])
                / comparison_df["Önceki Genel Toplam"] * 100
            ).round(2)
            comparison_df["Sert Düşüş"] = comparison_df["Değişim %"] <= (-drop_threshold)
            sharp_drops = comparison_df[comparison_df["Sert Düşüş"]].sort_values("Değişim %")

            if sharp_drops.empty:
                st.success("Sert düşüş eşiğini aşan anomali tespit edilmedi.")
            else:
                st.warning(f"{len(sharp_drops)} kalemde geçmiş döneme göre sert düşüş tespit edildi.")
                st.dataframe(sharp_drops, use_container_width=True, hide_index=True, height=360)
    else:
        st.info("Geçmiş dönem kıyaslaması için bu cihaz modunda hafıza verisi henüz yok.")

    render_consistency_section_banner(
        "Teklif Hafizasina Kaydet",
        "Bu modulu ileride daha akilli hale getirmek icin mevcut teklifi secilen varyant yapisiyla hafizaya yazin.",
        tone="slate",
        eyebrow="Kayit Merkezi",
    )
    st.caption("Kaydedilen kayıtlar, ileride otomatik kıyaslamalarda cihaz dahil/hariç ayrı referans olarak ve birim fiyat geçmiş havuzunda kullanılacaktır.")

    save_variant_answer_options = ["Hayır"]
    if current_offer_has_cihazlar:
        save_variant_answer_options = ["Her İkisi", "Evet", "Hayır"]

    save_variant_answer = st.radio(
        "Bu teklifte CİHAZLAR analize dahil mi?",
        options=save_variant_answer_options,
        horizontal=True,
        key="save_offer_memory_cihaz_mode",
    )

    if st.button("Bu Teklifi Hafızaya Kaydet", key="save_offer_memory"):
        if not teklif_kodu:
            st.error("Lütfen geçerli bir Teklif Kodu girin.")
        else:
            if save_variant_answer == "Evet":
                variants_to_save = [VARIANT_DEVICE_INCLUDED]
            elif save_variant_answer == "Her İkisi":
                variants_to_save = [VARIANT_DEVICE_INCLUDED, VARIANT_DEVICE_EXCLUDED]
            else:
                variants_to_save = [VARIANT_DEVICE_EXCLUDED]

            group_frames_to_save = [
                current_group_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_group_variants
            ]
            item_frames_to_save = [
                current_item_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_item_variants
            ]
            unit_price_frames_to_save = [
                current_unit_price_variants[variant_name]
                for variant_name in variants_to_save
                if variant_name in current_unit_price_variants
            ]
            group_memory_payload = pd.concat(group_frames_to_save, ignore_index=True) if group_frames_to_save else pd.DataFrame()
            item_memory_payload = pd.concat(item_frames_to_save, ignore_index=True) if item_frames_to_save else pd.DataFrame()
            unit_price_memory_payload = pd.concat(unit_price_frames_to_save, ignore_index=True) if unit_price_frames_to_save else pd.DataFrame()

            if group_memory_payload.empty and item_memory_payload.empty and unit_price_memory_payload.empty:
                st.error("Seçilen CİHAZLAR modu için kaydedilecek analiz verisi bulunamadı.")
            else:
                try:
                    if not group_memory_payload.empty:
                        upsert_memory_dataframe(
                            group_memory_path,
                            group_memory_payload,
                            key_columns=[MEMORY_VARIANT_KEY_COL, CONS_COL_GROUP],
                        )
                    if not item_memory_payload.empty:
                        upsert_memory_dataframe(
                            item_memory_path,
                            item_memory_payload,
                            key_columns=[MEMORY_VARIANT_KEY_COL, CONS_COL_GROUP, CONS_COL_PAGE, CONS_COL_DESC],
                        )
                    if not unit_price_memory_payload.empty:
                        upsert_memory_dataframe(
                            unit_price_memory_path,
                            unit_price_memory_payload,
                            key_columns=[MEMORY_VARIANT_KEY_COL, CONS_COL_UNIT_SCOPE, CONS_COL_GROUP, CONS_COL_UNIT],
                        )
                    st.success("Teklif hafızaya kaydedildi: " + ", ".join(variants_to_save))
                except Exception as save_error:
                    st.error(f"Hafıza kaydı sırasında hata: {str(save_error)}")

def create_sankey_chart(dataframe, label_col, value_col, source_label, title, max_targets=20, height=450):
    """Create a compact single-source Sankey chart from label/value columns."""
    sankey_df = dataframe[[label_col, value_col]].copy()
    sankey_df[value_col] = pd.to_numeric(sankey_df[value_col], errors='coerce').fillna(0)
    sankey_df = sankey_df[sankey_df[value_col] > 0]

    if sankey_df.empty:
        return None

    sankey_df[label_col] = sankey_df[label_col].astype(str).str.strip()
    sankey_df = sankey_df[sankey_df[label_col] != ""]
    sankey_df = sankey_df[sankey_df[label_col].str.lower() != "nan"]
    sankey_df = sankey_df.groupby(label_col, as_index=False)[value_col].sum()
    sankey_df = sankey_df.sort_values(value_col, ascending=False)

    if len(sankey_df) > max_targets:
        top_df = sankey_df.head(max_targets).copy()
        other_total = sankey_df.iloc[max_targets:][value_col].sum()
        if other_total > 0:
            top_df.loc[len(top_df)] = {label_col: "Diğer", value_col: other_total}
        sankey_df = top_df

    labels = [source_label] + sankey_df[label_col].tolist()
    sources = [0] * len(sankey_df)
    targets = list(range(1, len(sankey_df) + 1))
    values = sankey_df[value_col].tolist()
    color_palette = [
        "#2563eb", "#16a34a", "#f59e0b", "#ef4444", "#8b5cf6",
        "#06b6d4", "#ec4899", "#84cc16", "#f97316", "#14b8a6",
    ]
    target_node_colors = [color_palette[i % len(color_palette)] for i in range(len(sankey_df))]

    def hex_to_rgba(hex_color, alpha=0.45):
        hex_clean = hex_color.lstrip("#")
        if len(hex_clean) != 6:
            return f"rgba(59,130,246,{alpha})"
        red = int(hex_clean[0:2], 16)
        green = int(hex_clean[2:4], 16)
        blue = int(hex_clean[4:6], 16)
        return f"rgba({red},{green},{blue},{alpha})"

    link_colors = [hex_to_rgba(color, 0.38) for color in target_node_colors]

    fig = go.Figure(
        data=[
            go.Sankey(
                arrangement="snap",
                textfont=dict(
                    family="Arial, sans-serif",
                    size=14,
                    color="#000000",
                ),
                node=dict(
                    pad=15,
                    thickness=18,
                    line=dict(color="rgba(0,0,0,0)", width=0),
                    label=labels,
                    color=["#1d4ed8"] + target_node_colors,
                ),
                link=dict(
                    source=sources,
                    target=targets,
                    value=values,
                    color=link_colors,
                ),
            )
        ]
    )
    fig.update_layout(
        title=dict(text=title, font=dict(family="Arial, sans-serif", size=18, color="#000000")),
        height=height,
        font=dict(family="Arial, sans-serif", size=13, color="#000000"),
        hoverlabel=dict(
            font=dict(family="Arial, sans-serif", size=12, color="#000000"),
            bgcolor="#ffffff",
            bordercolor="rgba(0,0,0,0)"
        ),
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        margin=dict(l=20, r=20, t=60, b=20),
    )
    return fig

def create_sortable_numeric_table(dataframe, height=400, currency_cols=None, percent_cols=None, integer_cols=None):
    """Render sortable numeric tables with the same visual style as analysis tables."""
    currency_cols = set(currency_cols or [])
    percent_cols = set(percent_cols or [])
    integer_cols = set(integer_cols or [])

    display_df = dataframe.copy()
    display_df = move_action_columns_to_front(display_df)
    for col in currency_cols.union(percent_cols).union(integer_cols):
        if col in display_df.columns:
            display_df[col] = pd.to_numeric(display_df[col], errors="coerce")

    if any(is_action_column_name(col) for col in display_df.columns):
        return create_aggrid_table(
            display_df,
            height=height,
            selection_mode=None,
            fit_columns_on_grid_load=False,
            currency_cols=list(currency_cols),
            percent_cols=list(percent_cols),
            integer_cols=list(integer_cols),
        )

    def apply_column_styles(styler):
        malzeme_cols = []
        iscilik_cols = []
        ggk_cols = []
        toplam_cols = []
        grup_cols = []

        for col in display_df.columns:
            normalized = normalize_excel_header_name(col)
            if "MALZEME" in normalized:
                malzeme_cols.append(col)
            if "ISCILIK" in normalized:
                iscilik_cols.append(col)
            if "GGK" in normalized:
                ggk_cols.append(col)
            if (
                "GENEL TOPLAM" in normalized
                or "KUMULATIF GENEL" in normalized
                or "GENEL DEGISIM" in normalized
                or "GENEL INDIRIM" in normalized
                or "GENEL ZAM" in normalized
            ):
                toplam_cols.append(col)
            if "URUN GRUBU" in normalized:
                grup_cols.append(col)

        for col in malzeme_cols:
            styler = styler.map(lambda _: "background-color: #dcfce7; color: #15803d; font-weight: 700", subset=[col])
        for col in iscilik_cols:
            styler = styler.map(lambda _: "background-color: #dbeafe; color: #1d4ed8; font-weight: 700", subset=[col])
        for col in ggk_cols:
            styler = styler.map(lambda _: "background-color: #e9d5ff; color: #7c3aed; font-weight: 700", subset=[col])
        for col in toplam_cols:
            styler = styler.map(lambda _: "background-color: #fecaca; color: #dc2626; font-weight: 800; font-size: 15px", subset=[col])
        for col in grup_cols:
            styler = styler.map(lambda _: "background-color: #f3f4f6; color: #374151; font-weight: 800; font-size: 13px", subset=[col])

        # Satır bazında İndirim/Zam vurgusu: detay tabloda Durum'a göre, özet tabloda fark işaretine göre.
        status_target_cols = [
            col
            for col in [
                "Durum",
                "Genel Toplam Farkı",
                "Genel Toplam Farkı %",
                "Genel Değişim %",
                "Değişim %",
                "Genel İndirim Tutarı",
                "İndirim Tutarı",
                "Genel Zam Tutarı",
                "Zam Tutarı",
            ]
            if col in display_df.columns
        ]
        if "Durum" in display_df.columns and status_target_cols:
            def style_status_row(row):
                status_text = turkce_ascii(str(row.get("Durum", ""))).upper().strip()
                if status_text == "INDIRIM":
                    style = "color: #15803d; font-weight: 800"
                elif status_text == "ZAM":
                    style = "color: #dc2626; font-weight: 800"
                else:
                    style = ""
                return [style for _ in status_target_cols]

            styler = styler.apply(style_status_row, axis=1, subset=status_target_cols)
        elif "Genel Toplam Farkı" in display_df.columns:
            summary_target_cols = [
                col
                for col in [
                    "Genel Toplam Farkı",
                    "Genel Toplam Farkı %",
                    "Genel Değişim %",
                    "Değişim %",
                    "Genel İndirim Tutarı",
                    "İndirim Tutarı",
                    "Genel Zam Tutarı",
                    "Zam Tutarı",
                ]
                if col in display_df.columns
            ]
            if summary_target_cols:
                def style_change_row(row):
                    fark = pd.to_numeric(row.get("Genel Toplam Farkı"), errors="coerce")
                    if pd.isna(fark) or fark == 0:
                        style = ""
                    elif fark < 0:
                        style = "color: #15803d; font-weight: 800"
                    else:
                        style = "color: #dc2626; font-weight: 800"
                    return [style for _ in summary_target_cols]

                styler = styler.apply(style_change_row, axis=1, subset=summary_target_cols)

        return styler

    st.markdown("""
    <style>
    div[data-testid="stDataFrame"] {
        border: 2px solid #000000 !important;
        border-radius: 16px !important;
        overflow: hidden !important;
        box-shadow: 0 8px 24px rgba(0,0,0,0.12) !important;
        font-family: "Aptos Narrow", Aptos, "Segoe UI", sans-serif !important;
    }
    div[data-testid="stDataFrame"] th {
        background: linear-gradient(135deg, #7c3aed 0%, #a855f7 50%, #c084fc 100%) !important;
        color: #ffffff !important;
        font-weight: 900 !important;
        font-size: 10px !important;
        text-transform: uppercase !important;
        letter-spacing: 0.35px !important;
        padding: 17px 18px !important;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3) !important;
        text-align: left !important;
        line-height: 1.15 !important;
        font-family: "Aptos Narrow", Aptos, "Segoe UI", sans-serif !important;
        border-right: 1.5px solid #000000 !important;
        border-bottom: 2px solid #000000 !important;
    }
    div[data-testid="stDataFrame"] [role="columnheader"] {
        background: linear-gradient(135deg, #7c3aed 0%, #a855f7 50%, #c084fc 100%) !important;
        color: #ffffff !important;
        font-weight: 900 !important;
        font-size: 10px !important;
        letter-spacing: 0.35px !important;
        text-align: left !important;
        border-bottom: 2px solid #5b21b6 !important;
        font-family: "Aptos Narrow", Aptos, "Segoe UI", sans-serif !important;
    }
    div[data-testid="stDataFrame"] [role="columnheader"] * {
        color: #ffffff !important;
        font-weight: 900 !important;
        font-size: 10px !important;
        text-transform: uppercase !important;
        line-height: 1.15 !important;
        font-family: "Aptos Narrow", Aptos, "Segoe UI", sans-serif !important;
    }
    div[data-testid="stDataFrame"] td {
        font-weight: 600 !important;
        font-size: 16px !important;
        padding: 10px 16px !important;
        text-align: left !important;
        font-family: "Aptos Narrow", Aptos, "Segoe UI", sans-serif !important;
        border-right: 1.5px solid #000000 !important;
        border-bottom: 1.5px solid #000000 !important;
    }
    @media print {
        div[data-testid="stDataFrame"] {
            overflow: visible !important;
            border-radius: 0 !important;
            box-shadow: none !important;
            width: 100% !important;
            max-width: 100% !important;
            page-break-inside: avoid !important;
            break-inside: avoid !important;
        }
        div[data-testid="stDataFrame"] * {
            max-height: none !important;
        }
        div[data-testid="stDataFrame"] table {
            width: 100% !important;
            table-layout: auto !important;
        }
        div[data-testid="stDataFrame"] th,
        div[data-testid="stDataFrame"] td {
            font-size: 10px !important;
            padding: 4px 6px !important;
            white-space: normal !important;
            word-break: break-word !important;
            line-height: 1.2 !important;
        }
    }
    </style>
    """, unsafe_allow_html=True)

    styler = display_df.style.pipe(apply_column_styles)
    formatters = {}
    for col in currency_cols:
        if col in display_df.columns:
            formatters[col] = format_currency_display
    for col in percent_cols:
        if col in display_df.columns:
            formatters[col] = format_percent_display
    for col in integer_cols:
        if col in display_df.columns:
            formatters[col] = format_integer_display
    if formatters:
        styler = styler.format(formatters)

    return st.dataframe(
        styler,
        height=height,
        use_container_width=True,
        hide_index=True,
    )

def normalize_grid_column_name(value):
    """Normalize grid column names for pinning and action-column detection."""
    text = "" if value is None else str(value)
    replacements = {
        "\u0131": "i", "\u011f": "g", "\u00fc": "u", "\u015f": "s", "\u00f6": "o", "\u00e7": "c",
        "\u0130": "I", "\u011e": "G", "\u00dc": "U", "\u015e": "S", "\u00d6": "O", "\u00c7": "C",
        "ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c",
        "İ": "I", "Ğ": "G", "Ü": "U", "Ş": "S", "Ö": "O", "Ç": "C",
        "Ä±": "i", "ÄŸ": "g", "Ã¼": "u", "ÅŸ": "s", "Ã¶": "o", "Ã§": "c",
        "Ä°": "I", "Ä": "G", "Ãœ": "U", "Å": "S", "Ã–": "O", "Ã‡": "C",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text.upper().strip()

def is_action_column_name(col_name):
    normalized = normalize_grid_column_name(col_name).replace(" ", "")
    action_tokens = ["ACTION", "ACTIONS", "AKSIYON", "AKSIYONLAR", "ISLEM", "ISLEMLER"]
    return any(token in normalized for token in action_tokens)

def move_action_columns_to_front(dataframe):
    if dataframe is None:
        return pd.DataFrame()

    display_df = dataframe.copy()
    action_columns = [col for col in display_df.columns if is_action_column_name(col)]
    if not action_columns:
        return display_df

    remaining_columns = list(display_df.columns)
    for action_column in action_columns:
        if action_column in remaining_columns:
            remaining_columns.remove(action_column)
    ordered_columns = action_columns + remaining_columns
    return display_df.loc[:, ordered_columns]

def create_aggrid_table(
    dataframe,
    height=400,
    selection_mode='single',
    fit_columns_on_grid_load=True,
    currency_cols=None,
    percent_cols=None,
    integer_cols=None,
):
    """Render dataframe table with preserved column order and manually resizable columns."""

    explicit_currency_cols = {str(col) for col in (currency_cols or [])}
    explicit_percent_cols = {str(col) for col in (percent_cols or [])}
    explicit_integer_cols = {str(col) for col in (integer_cols or [])}

    def is_currency_column(col_name):
        text = str(col_name)
        if text in explicit_currency_cols:
            return True
        if text in explicit_percent_cols:
            return False
        if text in explicit_integer_cols:
            return False
        if "%" in text:
            return False
        if text in PRICE_COLUMNS:
            return True

        normalized = normalize_excel_header_name(text)
        exclude_tokens = ["SAYISI", "SATIR", "SAYFA", "SIRA", "TARIH", "KODU"]
        if any(token in normalized for token in exclude_tokens):
            return False

        currency_tokens = ["FIYATI", "GENEL TOPLAM", "TUTARI", "MALIYET", "FARKI", "NET FARK"]
        return any(token in normalized for token in currency_tokens)

    def estimate_column_width(df, col_name):
        normalized = normalize_excel_header_name(col_name)
        series = df[col_name] if col_name in df.columns else pd.Series(dtype="object")
        sample_lengths = series.dropna().astype(str).head(80).map(len)
        sample_length = int(sample_lengths.quantile(0.85)) if not sample_lengths.empty else 0
        content_length = max(len(str(col_name)), sample_length)

        if is_action_column_name(col_name):
            min_width, max_width, char_px = 120, 190, 8
        elif "URUN GRUBU" in normalized:
            min_width, max_width, char_px = 150, 240, 8
        elif "ACIKLAMA" in normalized:
            min_width, max_width, char_px = 160, 260, 8
        elif is_currency_column(col_name):
            min_width, max_width, char_px = 110, 160, 7
        elif "%" in str(col_name):
            min_width, max_width, char_px = 90, 120, 7
        elif any(token in normalized for token in ["SAYISI", "SATIR", "SIRA"]):
            min_width, max_width, char_px = 80, 110, 7
        else:
            min_width, max_width, char_px = 100, 150, 7

        estimated_width = 28 + (content_length * char_px)
        return max(min_width, min(max_width, estimated_width))

    display_df = dataframe.copy() if dataframe is not None else pd.DataFrame()

    try:
        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

        gb = GridOptionsBuilder.from_dataframe(display_df)
        gb.configure_default_column(
            resizable=True,
            sortable=True,
            filter=True,
            minWidth=90,
            wrapHeaderText=True,
            autoHeaderHeight=True,
        )
        if selection_mode in {"single", "multiple"}:
            gb.configure_selection(selection_mode=selection_mode, use_checkbox=False)

        gb.configure_grid_options(
            headerHeight=42,
            rowHeight=36,
            suppressMovableColumns=True,
            suppressDragLeaveHidesColumns=True,
            maintainColumnOrder=True,
            ensureDomOrder=True,
            animateRows=False,
        )

        currency_formatter = JsCode(
            """
            function(params) {
                if (params.value === null || params.value === undefined || params.value === "") return "";
                const value = Number(params.value);
                if (Number.isNaN(value)) return params.value;
                return value.toLocaleString("tr-TR", { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + " TL";
            }
            """
        )
        percent_formatter = JsCode(
            """
            function(params) {
                if (params.value === null || params.value === undefined || params.value === "") return "";
                const value = Number(params.value);
                if (Number.isNaN(value)) return params.value;
                return value.toLocaleString("tr-TR", { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + "%";
            }
            """
        )
        integer_formatter = JsCode(
            """
            function(params) {
                if (params.value === null || params.value === undefined || params.value === "") return "";
                const value = Number(params.value);
                if (Number.isNaN(value)) return params.value;
                return value.toLocaleString("tr-TR", { maximumFractionDigits: 0 });
            }
            """
        )

        for col in display_df.columns:
            col_name = str(col)
            normalized = normalize_excel_header_name(col_name)
            header_class = "grid-default-header"
            cell_style = {
                "textAlign": "left",
                "fontWeight": "600",
                "fontFamily": "Aptos Narrow, Aptos, Segoe UI, sans-serif",
                "justifyContent": "flex-start",
            }
            base_width = estimate_column_width(display_df, col_name)
            resize_min_width = max(80, min(base_width, int(base_width * 0.72)))
            col_kwargs = {
                "width": base_width,
                "minWidth": resize_min_width,
                "suppressMovable": True,
            }

            if is_action_column_name(col_name):
                header_class = "slate-header"
                action_width = max(120, min(160, base_width))
                cell_style.update(
                    {
                        "backgroundColor": "#f8fafc",
                        "color": "#0f172a",
                        "fontWeight": "700",
                        "textAlign": "center",
                        "justifyContent": "center",
                    }
                )
                col_kwargs.update(
                    {
                        "pinned": "left",
                        "lockPinned": True,
                        "minWidth": 110,
                        "width": action_width,
                        "maxWidth": 240,
                        "suppressMovable": True,
                    }
                )
            elif "MALZEME" in normalized:
                header_class = "malzeme-header"
                cell_style.update({"backgroundColor": "#dcfce7", "color": "#15803d", "fontWeight": "700"})
            elif "ISCILIK" in normalized:
                header_class = "iscilik-header"
                cell_style.update({"backgroundColor": "#dbeafe", "color": "#1d4ed8", "fontWeight": "700"})
            elif "GGK" in normalized:
                header_class = "ggk-header"
                cell_style.update({"backgroundColor": "#e9d5ff", "color": "#7c3aed", "fontWeight": "700"})
            elif "GENEL TOPLAM" in normalized or "KUMULATIF GENEL" in normalized:
                header_class = "toplam-header"
                cell_style.update({"backgroundColor": "#fecaca", "color": "#dc2626", "fontWeight": "800"})
            elif "URUN GRUBU" in normalized:
                header_class = "grup-header"
                cell_style.update({"backgroundColor": "#f3f4f6", "color": "#374151", "fontWeight": "800", "fontSize": "13px"})
                col_kwargs["pinned"] = "left"
                col_kwargs["minWidth"] = 150
            elif "SAYISI" in normalized:
                header_class = "slate-header"

            if col_name in explicit_currency_cols or is_currency_column(col_name):
                col_kwargs["type"] = ["numericColumn", "numberColumnFilter"]
                col_kwargs["valueFormatter"] = currency_formatter
            elif col_name in explicit_percent_cols or "%" in col_name:
                col_kwargs["type"] = ["numericColumn", "numberColumnFilter"]
                col_kwargs["valueFormatter"] = percent_formatter
            elif col_name in explicit_integer_cols or any(token in normalized for token in ["SAYISI", "SATIR", "SIRA"]):
                col_kwargs["type"] = ["numericColumn", "numberColumnFilter"]
                col_kwargs["valueFormatter"] = integer_formatter

            gb.configure_column(
                col_name,
                headerClass=header_class,
                cellStyle=cell_style,
                **col_kwargs,
            )

        grid_options = gb.build()
        custom_css = {
            ".ag-root-wrapper": {
                "border": "2px solid #000000",
                "border-radius": "16px",
                "overflow": "hidden",
                "box-shadow": "0 8px 24px rgba(0,0,0,0.12)",
                "font-family": "\"Aptos Narrow\", Aptos, \"Segoe UI\", sans-serif",
            },
            ".ag-header": {
                "background": "linear-gradient(135deg, #7c3aed 0%, #a855f7 50%, #c084fc 100%)",
                "border-bottom": "2px solid #000000",
            },
            ".ag-header-row": {
                "background": "linear-gradient(135deg, #7c3aed 0%, #a855f7 50%, #c084fc 100%)",
            },
            ".ag-header-cell": {
                "background": "transparent",
                "border-right": "1.5px solid #000000",
                "font-family": "\"Aptos Narrow\", Aptos, \"Segoe UI\", sans-serif",
            },
            ".ag-header-cell-label": {
                "justify-content": "flex-start",
            },
            ".ag-header-cell-text": {
                "color": "#ffffff !important",
                "font-size": "10px !important",
                "font-weight": "900 !important",
                "text-transform": "uppercase",
                "letter-spacing": "0.35px",
                "line-height": "1.15",
                "font-family": "\"Aptos Narrow\", Aptos, \"Segoe UI\", sans-serif",
                "text-align": "left",
            },
            ".ag-cell": {
                "font-size": "16px",
                "font-weight": "600",
                "border-right": "1.5px solid #000000",
                "border-bottom": "1.5px solid #000000",
                "display": "flex",
                "align-items": "center",
                "justify-content": "flex-start",
                "font-family": "\"Aptos Narrow\", Aptos, \"Segoe UI\", sans-serif",
            },
            ".ag-cell-value": {
                "text-align": "left",
                "width": "100%",
            },
            ".ag-row-hover .ag-cell": {
                "background-color": "#faf5ff !important",
            },
            ".grup-header .ag-header-cell-text": {
                "font-weight": "900 !important",
            },
            ".toplam-header .ag-header-cell-text": {
                "font-weight": "900 !important",
            },
        }

        return AgGrid(
            display_df,
            gridOptions=grid_options,
            height=height,
            fit_columns_on_grid_load=fit_columns_on_grid_load,
            allow_unsafe_jscode=True,
            custom_css=custom_css,
            theme="streamlit",
            enable_enterprise_modules=False,
            update_mode="NO_UPDATE",
            data_return_mode="AS_INPUT",
            reload_data=False,
        )
    except Exception as e:
        st.warning(f"AgGrid stilleri uygulanamadi, standart tabloya donuluyor: {str(e)}")
        return st.dataframe(display_df, height=height, use_container_width=True, hide_index=True)


HEADING_TONE_MAP = {
    "GENEL GORSELLESTIRMELER": {
        "background": "linear-gradient(135deg, #ecfeff, #ccfbf1)",
        "border": "#14b8a6",
        "accent": "linear-gradient(90deg, #14b8a6, #0ea5e9)",
        "title": "#134e4a",
    },
    "URUN GRUPLARI DETAY ANALIZI": {
        "background": "linear-gradient(135deg, #fff1f2, #ffe4e6)",
        "border": "#f43f5e",
        "accent": "linear-gradient(90deg, #f43f5e, #fb7185)",
        "title": "#9f1239",
    },
    "KATEGORIYE GORE GIDER DAGILIMI": {
        "background": "linear-gradient(135deg, #eff6ff, #dbeafe)",
        "border": "#3b82f6",
        "accent": "linear-gradient(90deg, #2563eb, #38bdf8)",
        "title": "#1e3a8a",
    },
    "GIDER KALEMINE GORE DETAYLI ANALIZ": {
        "background": "linear-gradient(135deg, #fff7ed, #fed7aa)",
        "border": "#f97316",
        "accent": "linear-gradient(90deg, #f97316, #fb923c)",
        "title": "#9a3412",
    },
    "FIYAT REVIZYON KIYAS": {
        "background": "linear-gradient(135deg, #eef2ff, #e0e7ff)",
        "border": "#6366f1",
        "accent": "linear-gradient(90deg, #6366f1, #8b5cf6)",
        "title": "#3730a3",
    },
    "TUTARLILIK VE GECMIS KIYAS MODULU": {
        "background": "linear-gradient(135deg, #f0fdfa, #ccfbf1)",
        "border": "#0f766e",
        "accent": "linear-gradient(90deg, #0f766e, #14b8a6)",
        "title": "#134e4a",
    },
}

SUBHEADING_TONE_MAP = {
    "ABC ANALIZI": {
        "background": "linear-gradient(135deg, #fef3c7, #fde68a)",
        "border": "#f59e0b",
        "title": "#92400e",
    },
    "EN YUKSEK MALIYETLI ILK 20 URUN": {
        "background": "linear-gradient(135deg, #fee2e2, #fecaca)",
        "border": "#ef4444",
        "title": "#991b1b",
    },
    "GRUP ICI URUN DETAYLARI": {
        "background": "linear-gradient(135deg, #ede9fe, #ddd6fe)",
        "border": "#8b5cf6",
        "title": "#5b21b6",
    },
    "GORSELLESTIRMELER": {
        "background": "linear-gradient(135deg, #dcfce7, #bbf7d0)",
        "border": "#22c55e",
        "title": "#166534",
    },
    "MINIMUM VE MAXIMUM DEGERLER": {
        "background": "linear-gradient(135deg, #fce7f3, #fbcfe8)",
        "border": "#ec4899",
        "title": "#9d174d",
    },
    "PARETO ANALIZI": {
        "background": "linear-gradient(135deg, #cffafe, #bae6fd)",
        "border": "#06b6d4",
        "title": "#155e75",
    },
    "GIDER PARETO ANALIZI": {
        "background": "linear-gradient(135deg, #cffafe, #bae6fd)",
        "border": "#06b6d4",
        "title": "#155e75",
    },
}

def _resolve_heading_tone(title: str, subsection: bool = False):
    normalized_title = normalize_excel_header_name(title)
    tone_map = SUBHEADING_TONE_MAP if subsection else HEADING_TONE_MAP
    for key, tone in tone_map.items():
        if key in normalized_title:
            return tone
    return (
        {
            "background": "linear-gradient(135deg, #f8fafc, #e2e8f0)",
            "border": "#2563eb",
            "accent": "linear-gradient(90deg, #2563eb, #7c3aed)",
            "title": "#1e293b",
        }
        if not subsection
        else {
            "background": "linear-gradient(135deg, #fff7ed, #ffedd5)",
            "border": "#f59e0b",
            "title": "#7c2d12",
        }
    )

def render_section_heading(title: str, icon: str = "") -> None:
    display_title = f"{icon} {title}".strip() if icon else title
    tone = _resolve_heading_tone(display_title, subsection=False)
    st.markdown(
        f"""
        <div style="
            margin: 34px 0 16px 0;
            padding: 22px 24px;
            border-radius: 18px;
            background: {tone['background']};
            border: 1px solid {tone['border']};
            box-shadow: 0 14px 30px rgba(15,23,42,0.06);
            position: relative;
            overflow: hidden;
        ">
            <div style="
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 4px;
                background: {tone['accent']};
            "></div>
            <div style="
                font-size: 26px;
                font-weight: 900;
                letter-spacing: -0.02em;
                color: {tone['title']};
                line-height: 1.1;
            ">{escape(display_title)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_subsection_heading(title: str, icon: str = "") -> None:
    display_title = f"{icon} {title}".strip() if icon else title
    tone = _resolve_heading_tone(display_title, subsection=True)
    st.markdown(
        f"""
        <div style="
            margin: 24px 0 12px 0;
            padding: 14px 18px;
            border-radius: 14px;
            background: {tone['background']};
            border-left: 5px solid {tone['border']};
            box-shadow: 0 10px 22px rgba(15,23,42,0.05);
        ">
            <div style="
                font-size: 19px;
                font-weight: 800;
                color: {tone['title']};
                letter-spacing: -0.01em;
                line-height: 1.2;
            ">{escape(display_title)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def turkce_ascii(text):
    """Türkçe karakterleri ASCII'ye çevir"""
    if not isinstance(text, str):
        return str(text)
    tr_chars = {'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c',
                'İ': 'I', 'Ğ': 'G', 'Ü': 'U', 'Ş': 'S', 'Ö': 'O', 'Ç': 'C'}
    for tr, en in tr_chars.items():
        text = text.replace(tr, en)
    return text

def generate_pdf_report(combined_df, selected_sheets):
    """
    PDF raporu oluşturur - tüm ürün gruplarının detaylı analizi
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1.5*cm, bottomMargin=1.5*cm)

    # PDF elemanları
    elements = []

    # Başlık tablosu
    title_data = [['URUN GRUPLARINA GORE MALIYET ANALIZI RAPORU']]
    title_table = Table(title_data, colWidths=[26*cm])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 18),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(title_table)

    # Tarih
    date_data = [[f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"]]
    date_table = Table(date_data, colWidths=[26*cm])
    date_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(date_table)
    elements.append(Spacer(1, 0.5*cm))

    # Genel Özet başlık
    section_data = [['GENEL OZET']]
    section_table = Table(section_data, colWidths=[26*cm])
    section_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(section_table)
    elements.append(Spacer(1, 0.3*cm))

    total_malzeme = combined_df['Malzeme Fiyatı'].sum()
    total_iscilik = combined_df['İşçilik Fiyatı'].sum()
    total_ggk = combined_df['GGK Fiyatı'].sum()
    total_genel = combined_df['Genel Toplam'].sum()

    summary_data = [
        ['Kategori', 'Tutar (TL)', 'Yuzde (%)'],
        ['Toplam Malzeme', f'{total_malzeme:,.2f}', f'{(total_malzeme/total_genel*100):.2f}%'],
        ['Toplam Iscilik', f'{total_iscilik:,.2f}', f'{(total_iscilik/total_genel*100):.2f}%'],
        ['Toplam GGK', f'{total_ggk:,.2f}', f'{(total_ggk/total_genel*100):.2f}%'],
        ['GENEL TOPLAM', f'{total_genel:,.2f}', '100.00%']
    ]

    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fecaca')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#dc2626')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 1*cm))

    # Ürün Gruplarına Göre Detaylı Analiz
    section_data2 = [['URUN GRUPLARINA GORE DETAYLI ANALIZ']]
    section_table2 = Table(section_data2, colWidths=[26*cm])
    section_table2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(section_table2)
    elements.append(Spacer(1, 0.3*cm))

    # Tüm verileri gruplandır
    grouped = combined_df.groupby('Ürün Grubu').agg({
        'Malzeme Fiyatı': 'sum',
        'İşçilik Fiyatı': 'sum',
        'GGK Fiyatı': 'sum',
        'Genel Toplam': 'sum',
        'Sayfa': 'count'  # Kayıt sayısını say
    }).reset_index()

    # Sütun adını değiştir
    grouped.rename(columns={'Sayfa': 'Bulunan Kayıt Sayısı'}, inplace=True)

    grouped = grouped.sort_values('Genel Toplam', ascending=False)
    grouped['Genel Toplam %'] = (grouped['Genel Toplam'] / grouped['Genel Toplam'].sum() * 100)
    grouped['Kümülatif Genel %'] = grouped['Genel Toplam %'].cumsum()

    # Ana tablo verileri - Kısa başlıklar
    main_table_data = [['Urun Grubu', 'G.Toplam', 'G.Top %', 'Kum %',
                        'Malzeme', 'Mlz %', 'Iscilik', 'Isc %',
                        'GGK', 'GGK %', 'Adet']]

    for _, row in grouped.iterrows():
        main_table_data.append([
            turkce_ascii(str(row['Ürün Grubu']))[:25],  # Kısa isim
            f"{row['Genel Toplam']:,.0f}",
            f"{row['Genel Toplam %']:.1f}%",
            f"{row['Kümülatif Genel %']:.1f}%",
            f"{row['Malzeme Fiyatı']:,.0f}",
            f"{(row['Malzeme Fiyatı']/row['Genel Toplam']*100):.1f}%",
            f"{row['İşçilik Fiyatı']:,.0f}",
            f"{(row['İşçilik Fiyatı']/row['Genel Toplam']*100):.1f}%",
            f"{row['GGK Fiyatı']:,.0f}",
            f"{(row['GGK Fiyatı']/row['Genel Toplam']*100):.1f}%",
            f"{int(row['Bulunan Kayıt Sayısı'])}"
        ])

    # Tablo genişlikleri - Daha dengeli dağılım
    col_widths = [4.2*cm, 2.4*cm, 1.9*cm, 1.8*cm, 2.4*cm, 1.6*cm, 2.4*cm, 1.6*cm, 2.4*cm, 1.6*cm, 1.8*cm]

    main_table = Table(main_table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle([
        # Başlık satırı
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('WORDWRAP', (0, 0), (-1, -1), True),
        # Genel Toplam sütunları vurgula
        ('BACKGROUND', (1, 1), (3, -1), colors.HexColor('#fecaca')),
        ('TEXTCOLOR', (1, 1), (3, -1), colors.HexColor('#dc2626')),
        ('FONTNAME', (1, 1), (3, -1), 'Helvetica-Bold'),
        # Malzeme sütunları
        ('BACKGROUND', (4, 1), (5, -1), colors.HexColor('#d1fae5')),
        ('TEXTCOLOR', (4, 1), (5, -1), colors.HexColor('#047857')),
        # İşçilik sütunları
        ('BACKGROUND', (6, 1), (7, -1), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (6, 1), (7, -1), colors.HexColor('#1e40af')),
        # GGK sütunları
        ('BACKGROUND', (8, 1), (9, -1), colors.HexColor('#e9d5ff')),
        ('TEXTCOLOR', (8, 1), (9, -1), colors.HexColor('#7c3aed')),
    ]))

    elements.append(main_table)
    elements.append(Spacer(1, 0.5*cm))

    # Tum teklif icin en pahali 50 urun
    top_products_title_data = [['EN PAHALI 50 URUN (GENEL TOPLAM)']]
    top_products_title_table = Table(top_products_title_data, colWidths=[26*cm])
    top_products_title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#b91c1c')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(top_products_title_table)
    elements.append(Spacer(1, 0.2*cm))

    top_products_df = combined_df.sort_values('Genel Toplam', ascending=False).head(50).copy()
    top_products_data = [['#', 'Urun Grubu', 'Urun Aciklamasi', 'Sayfa', 'Satir', 'Genel Toplam (TL)']]

    for rank, (_, row) in enumerate(top_products_df.iterrows(), start=1):
        top_products_data.append([
            str(rank),
            turkce_ascii(str(row.get('Ürün Grubu', '')))[:24],
            turkce_ascii(str(row.get('Ürün Açıklaması', '')))[:46],
            turkce_ascii(str(row.get('Sayfa', '')))[:16],
            str(row.get('Satır', '')),
            f"{float(row.get('Genel Toplam', 0) or 0):,.2f}",
        ])

    top_products_table = Table(
        top_products_data,
        colWidths=[1.1*cm, 4.0*cm, 9.1*cm, 3.0*cm, 2.0*cm, 5.0*cm],
        repeatRows=1
    )
    top_products_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7f1d1d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (5, 1), (5, -1), colors.HexColor('#b91c1c')),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    elements.append(top_products_table)
    elements.append(PageBreak())

    # Her ürün grubu için detaylı sayfa
    for _, group_row in grouped.iterrows():
        urun_grubu = group_row['Ürün Grubu']

        # Ürün grubu başlığı
        group_title_data = [[f"URUN GRUBU DETAYI: {turkce_ascii(str(urun_grubu))}"]]
        group_title_table = Table(group_title_data, colWidths=[26*cm])
        group_title_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(group_title_table)
        elements.append(Spacer(1, 0.3*cm))

        # Bu ürün grubuna ait tüm kayıtları getir
        group_df = combined_df[combined_df['Ürün Grubu'] == urun_grubu].copy()

        # Özet bilgiler
        group_summary_data = [
            ['Metrik', 'Deger'],
            ['Toplam Kayit Sayisi', f"{len(group_df)}"],
            ['Toplam Maliyet', f"{group_row['Genel Toplam']:,.2f} TL"],
            ['Malzeme Maliyeti', f"{group_row['Malzeme Fiyatı']:,.2f} TL ({(group_row['Malzeme Fiyatı']/group_row['Genel Toplam']*100):.1f}%)"],
            ['Iscilik Maliyeti', f"{group_row['İşçilik Fiyatı']:,.2f} TL ({(group_row['İşçilik Fiyatı']/group_row['Genel Toplam']*100):.1f}%)"],
            ['GGK Maliyeti', f"{group_row['GGK Fiyatı']:,.2f} TL ({(group_row['GGK Fiyatı']/group_row['Genel Toplam']*100):.1f}%)"],
        ]

        group_summary_table = Table(group_summary_data, colWidths=[8*cm, 10*cm])
        group_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(group_summary_table)
        elements.append(Spacer(1, 0.5*cm))

        # Detaylı kayıt tablosu (ilk 50 kayıt)
        group_df_sorted = group_df.sort_values('Genel Toplam', ascending=False).copy()
        detail_title_data = [[f"Detayli Kayitlar (En Pahali {min(50, len(group_df_sorted))} Kayit)"]]
        detail_title_table = Table(detail_title_data, colWidths=[26*cm])
        detail_title_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(detail_title_table)
        elements.append(Spacer(1, 0.2*cm))

        detail_data = [['Sayfa', 'Satir', 'Malzeme (TL)', 'Iscilik (TL)', 'GGK (TL)', 'Genel Toplam (TL)']]

        for idx, (_, row) in enumerate(group_df_sorted.head(50).iterrows()):
            detail_data.append([
                turkce_ascii(str(row.get('Sayfa', ''))),
                str(row.get('Satır', '')),
                f"{row['Malzeme Fiyatı']:,.2f}",
                f"{row['İşçilik Fiyatı']:,.2f}",
                f"{row['GGK Fiyatı']:,.2f}",
                f"{row['Genel Toplam']:,.2f}"
            ])

        detail_table = Table(detail_data, colWidths=[3*cm, 3*cm, 4*cm, 4*cm, 4*cm, 4.5*cm])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6b21a8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(detail_table)

        if len(group_df) > 50:
            elements.append(Spacer(1, 0.3*cm))
            note_data = [[f"Not: Bu urun grubunda toplam {len(group_df)} kayit bulunmaktadir. En pahali 50 kayit gosterilmistir."]]
            note_table = Table(note_data, colWidths=[26*cm])
            note_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Oblique'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
            ]))
            elements.append(note_table)

        elements.append(PageBreak())

    # PDF'i oluştur
    doc.build(elements)
    buffer.seek(0)
    return buffer

